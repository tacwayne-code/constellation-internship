import os, uuid
from datetime import datetime, date
from flask import Blueprint, render_template, redirect, url_for, request, flash, jsonify, current_app, session
from flask_login import login_user, logout_user, login_required, current_user
from werkzeug.utils import secure_filename
from app import db
from app.models import *

# ─── Blueprints ───
auth_bp = Blueprint('auth', __name__)
main_bp = Blueprint('main', __name__)
document_bp = Blueprint('documents', __name__)
workflow_bp = Blueprint('workflows', __name__)
structure_bp = Blueprint('structures', __name__)
bom_bp = Blueprint('boms', __name__)
project_bp = Blueprint('projects', __name__)
change_bp = Blueprint('changes', __name__)
workhour_bp = Blueprint('work_hours', __name__)
process_bp = Blueprint('processes', __name__)
integration_bp = Blueprint('integrations', __name__)


# ─── 辅助函数 ───
def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in \
        {'pdf', 'dwg', 'dxf', 'step', 'stp', 'igs', 'prt', 'asm', 'sldprt', 'sldasm',
         'doc', 'docx', 'xls', 'xlsx', 'ppt', 'pptx', 'txt', 'zip', 'rar', 'jpg', 'png'}


def require_role(*roles):
    """检查当前用户是否拥有指定角色之一，否则 abort(403)"""
    if not current_user.is_authenticated:
        from flask import abort as _abort
        _abort(401)
    if current_user.role not in roles:
        from flask import abort as _abort
        _abort(403)


def require_bom_write(bom=None):
    """BOM 写操作权限：admin/manager 可操作，普通用户只能操作自己创建的 BOM"""
    if not current_user.is_authenticated:
        from flask import abort as _abort
        _abort(401)
    if current_user.role in ('admin', 'manager'):
        return
    if bom is not None and bom.created_by == current_user.id:
        return
    from flask import abort as _abort
    _abort(403)


def save_file(file):
    if file and allowed_file(file.filename):
        ext = file.filename.rsplit('.', 1)[1].lower()
        fname = f"{uuid.uuid4().hex}.{ext}"
        folder = current_app.config['UPLOAD_FOLDER']
        os.makedirs(folder, exist_ok=True)
        path = os.path.join(folder, fname)
        file.save(path)
        return fname, file.filename, os.path.getsize(path)
    return None, None, 0


# ════════════════════════════════
#  AUTH
# ════════════════════════════════

# ── DB 路径（跨平台兼容，从环境变量或配置文件推断） ──
def _get_db_path():
    """获取 SQLite 数据库文件的绝对路径"""
    db_uri = current_app.config.get('SQLALCHEMY_DATABASE_URI', '')
    if db_uri.startswith('sqlite:///'):
        return db_uri.replace('sqlite:///', '')
    # 回退：从模块位置推断
    return os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))), 'plm.db')


@auth_bp.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        user = User.query.filter_by(username=request.form['username']).first()
        if user and user.check_password(request.form['password']) and user.is_active:
            login_user(user)
            return redirect(url_for('main.dashboard'))
        flash('用户名或密码错误', 'danger')
    return render_template('login.html')


@auth_bp.route('/logout')
@login_required
def logout():
    logout_user()
    return redirect(url_for('auth.login'))


# ════════════════════════════════
#  MAIN / DASHBOARD
# ════════════════════════════════
@main_bp.route('/')
@login_required
def dashboard():
    doc_count = Document.query.count()
    active_projects = Project.query.filter_by(status='active').count()
    pending_changes = ChangeRequest.query.filter(
        ChangeRequest.status.in_(['submitted', 'analyzing'])).count()

    # 待我审批的（文档 + BOM）
    my_pending_doc_approvals_q = DocApproval.query.filter_by(
        approver_id=current_user.id, status='pending')
    my_pending_approvals_total = my_pending_doc_approvals_q.count()
    my_pending_bom_approvals_q = BomApproval.query.filter_by(
        approver_id=current_user.id, status='pending')
    my_pending_bom_total = my_pending_bom_approvals_q.count()
    pending_approvals = my_pending_approvals_total + my_pending_bom_total

    # 我提交的文档审批进度
    my_submitted_doc_q = DocApproval.query.join(Document).filter(
        Document.author_id == current_user.id,
        DocApproval.status.in_(['pending', 'approved', 'rejected'])
    )
    my_submitted_doc_approvals_total = my_submitted_doc_q.count()

    # 我提交的 BOM 审批进度
    my_submitted_bom_q = BomApproval.query.filter(
        BomApproval.submitter_id == current_user.id
    )
    my_submitted_bom_approvals_total = my_submitted_bom_q.count()

    # 我审批过的记录（approver=我，已 approved/rejected）
    my_decided_bom_q = BomApproval.query.filter(
        BomApproval.approver_id == current_user.id,
        BomApproval.status.in_(['approved', 'rejected'])
    )
    my_decided_bom_approvals = my_decided_bom_q.order_by(
        BomApproval.decided_at.desc()
    ).limit(10).all()

    # 翻页参数
    docs_per_page = 5
    bom_per_page = 10
    submitted_per_page = 10

    docs_page = request.args.get('docs_page', 1, type=int)
    doc_approvals_page = request.args.get('doc_approvals_page', 1, type=int)
    bom_approvals_page = request.args.get('bom_approvals_page', 1, type=int)
    submitted_doc_page = request.args.get('submitted_doc_page', 1, type=int)
    submitted_bom_page = request.args.get('submitted_bom_page', 1, type=int)

    my_pending_approvals = my_pending_doc_approvals_q.order_by(
        DocApproval.created_at.desc()
    ).offset((doc_approvals_page - 1) * docs_per_page).limit(docs_per_page).all()
    my_pending_bom_approvals = my_pending_bom_approvals_q.order_by(
        BomApproval.created_at.desc()
    ).offset((bom_approvals_page - 1) * bom_per_page).limit(bom_per_page).all()
    my_submitted_doc_approvals = my_submitted_doc_q.order_by(
        DocApproval.created_at.desc()
    ).offset((submitted_doc_page - 1) * submitted_per_page).limit(submitted_per_page).all()
    my_submitted_bom_approvals = my_submitted_bom_q.order_by(
        BomApproval.created_at.desc()
    ).offset((submitted_bom_page - 1) * submitted_per_page).limit(submitted_per_page).all()

    # 计算总页数
    doc_approvals_pages = max(1, (my_pending_approvals_total + docs_per_page - 1) // docs_per_page)
    bom_approvals_pages = max(1, (my_pending_bom_total + bom_per_page - 1) // bom_per_page)
    submitted_doc_pages = max(1, (my_submitted_doc_approvals_total + submitted_per_page - 1) // submitted_per_page)
    submitted_bom_pages = max(1, (my_submitted_bom_approvals_total + submitted_per_page - 1) // submitted_per_page)

    # 我的待办变更
    my_pending_changes = ChangeRequest.query.filter(
        ChangeRequest.assignee_id == current_user.id,
        ChangeRequest.status.in_(['submitted', 'analyzing'])).limit(5).all()

    recent_docs = Document.query.order_by(Document.updated_at.desc()).limit(5).all()
    recent_changes = ChangeRequest.query.order_by(ChangeRequest.created_at.desc()).limit(5).all()

    # BOM 统计
    ebom_count = Bom.query.filter_by(bom_type='EBOM').count()
    mbom_count = Bom.query.filter_by(bom_type='MBOM').count()
    # 待推送的已发布 mBOM
    pending_sync_mboms = Bom.query.filter_by(
        bom_type='MBOM', status='released', sync_status='not_synced'
    ).count()

    return render_template('dashboard.html', doc_count=doc_count, pending_approvals=pending_approvals,
                           active_projects=active_projects, pending_changes=pending_changes,
                           my_pending_approvals=my_pending_approvals,
                           my_pending_bom_approvals=my_pending_bom_approvals,
                           my_submitted_doc_approvals=my_submitted_doc_approvals,
                           my_submitted_bom_approvals=my_submitted_bom_approvals,
                           my_decided_bom_approvals=my_decided_bom_approvals,
                           doc_approvals_page=doc_approvals_page, doc_approvals_pages=doc_approvals_pages,
                           docs_per_page=docs_per_page,
                           bom_approvals_page=bom_approvals_page, bom_approvals_pages=bom_approvals_pages,
                           bom_per_page=bom_per_page,
                           submitted_doc_page=submitted_doc_page, submitted_doc_pages=submitted_doc_pages,
                           submitted_bom_page=submitted_bom_page, submitted_bom_pages=submitted_bom_pages,
                           submitted_per_page=submitted_per_page,
                           my_pending_changes=my_pending_changes,
                           recent_docs=recent_docs, recent_changes=recent_changes,
                           ebom_count=ebom_count, mbom_count=mbom_count,
                           pending_sync_mboms=pending_sync_mboms)


# ════════════════════════════════
#  1. 图文档案管理
# ════════════════════════════════
@document_bp.route('/')
@login_required
def list_docs():
    category_id = request.args.get('category_id', type=int)
    keyword = request.args.get('q', '').strip()
    status_filter = request.args.get('status', '')
    fmt_filter = request.args.get('fmt', '').strip()  # PDF/CAD/3D 格式筛选
    page = request.args.get('page', 1, type=int)
    per_page = 20

    # 保存当前完整列表 URL 到 session，供详情页"返回列表"使用
    from flask import request as _req, session
    session['docs_list_url'] = _req.full_path.rstrip('?').replace('%2F', '/')
    # 如果没有任何筛选参数，标记为默认页（详情页可选默认行为）
    has_filters = bool(keyword or category_id or status_filter or fmt_filter or (page and page != 1))
    if not has_filters:
        session['docs_list_url'] = url_for('documents.list_docs')
    # 标记最近一次访问是列表
    session['docs_last_view'] = 'list'

    query = Document.query
    if category_id:
        query = query.filter_by(category_id=category_id)
    if keyword:
        like = f'%{keyword}%'
        # 搜索字段：doc_no（图号）+ title（图号）+ name（汉字名称）+ tags
        # 不查 description — 那是导入路径信息，噪音太大（如"导"字会命中"导入"路径）
        query = query.filter(db.or_(
            Document.title.ilike(like),
            Document.doc_no.ilike(like),
            Document.name.ilike(like),
            Document.tags.ilike(like)
        ))
    if status_filter:
        query = query.filter_by(status=status_filter)
    if fmt_filter:
        query = query.filter(Document.file_name.ilike(f'%.{fmt_filter}'))

    # 分页
    total = query.count()
    docs = query.order_by(Document.updated_at.desc()).offset((page - 1) * per_page).limit(per_page).all()

    categories = DocumentCategory.query.all()
    from sqlalchemy import func
    cat_rows = db.session.query(Document.category_id, func.count(Document.id)).group_by(Document.category_id).all()
    category_counts = {cid: cnt for cid, cnt in cat_rows if cid is not None}
    for cat in categories:
        if cat.id not in category_counts:
            category_counts[cat.id] = 0

    # 当前分类下各格式的文档数（用于子筛选 tab）
    fmt_counts = {}
    if category_id:
        # 用 LIKE 匹配完整扩展名（处理 .STEP .SLDDRW 这种长度 > 3 的）
        exts_to_count = ['pdf', 'dwg', 'drw', 'dxf', 'slddrw', 'sldprt', 'sldasm', 'step', 'stp', 'xt', 'doc', 'docx', 'xls', 'xlsx', 'zip', 'rar', 'jpg', 'png']
        for ext in exts_to_count:
            cnt = Document.query.filter(
                Document.category_id == category_id,
                Document.file_name.ilike(f'%.{ext}')
            ).count()
            if cnt > 0:
                fmt_counts[ext] = cnt

    total_pages = (total + per_page - 1) // per_page
    return render_template('documents/list.html', docs=docs, categories=categories,
                           category_counts=category_counts,
                           keyword=keyword, status_filter=status_filter,
                           fmt_filter=fmt_filter, fmt_counts=fmt_counts,
                           page=page, per_page=per_page, total=total, total_pages=total_pages,
                           category_id=category_id)


@document_bp.route('/create', methods=['GET', 'POST'])
@login_required
def create_doc():
    if request.method == 'POST':
        drawing_no = request.form.get('title', '').strip()
        name = request.form.get('name', '').strip()

        # 校验图号格式（8 位 XX-XX-XX-XX 或 9 位 XXX-XXX-XXX）
        import re
        if not (re.match(r'^\d{2}-\d{2}-\d{2}-\d{2}$', drawing_no) or
                re.match(r'^\d{3}-\d{3}-\d{3}$', drawing_no)):
            flash('图号格式错误：8 位 XX-XX-XX-XX 或 9 位 XXX-XXX-XXX', 'danger')
            categories = DocumentCategory.query.all()
            return render_template('documents/create.html', categories=categories)

        # 校验名称：必填，不允许包含图号
        if not name:
            flash('请填写图纸名称', 'danger')
            categories = DocumentCategory.query.all()
            return render_template('documents/create.html', categories=categories)
        if drawing_no in name:
            flash(f'名称中不允许包含图号「{drawing_no}」', 'danger')
            categories = DocumentCategory.query.all()
            return render_template('documents/create.html', categories=categories)

        doc = Document(
            title=drawing_no,           # title 存图号
            name=name,                  # name 存中文名称
            description=request.form.get('description', ''),
            category_id=request.form.get('category_id', type=int),
            tags=request.form.get('tags', ''),
            author_id=current_user.id
        )
        file = request.files.get('file')
        if file and file.filename:
            fname, orig_name, fsize = save_file(file)
            if fname:
                doc.file_name = orig_name
                doc.file_path = fname
                doc.file_size = fsize
        db.session.add(doc)
        db.session.commit()
        flash('文档创建成功', 'success')
        return redirect(url_for('documents.list_docs'))
    categories = DocumentCategory.query.all()
    return render_template('documents/create.html', categories=categories)


@document_bp.route('/tree')
@login_required
def tree_view():
    """树形浏览 — 按原图纸文件夹路径层级展开"""
    import re
    from collections import defaultdict
    bs = chr(92)
    selected_path = request.args.get('path', '').strip()
    keyword = request.args.get('q', '').strip()
    category_id = request.args.get('category_id', type=int)

    # 1) 收集所有路径 → 文档
    # 父目录映射：把已知的叶节点包装到机械图的正确层级下
    # 让树状浏览与原始文件夹层级保持一致
    FOLDER_PARENT = {
        # 上下板机
        '下板机打包2019 12 14': '上下板机',
        # 分光 编带机 > 常规生产
        '710分光机': '分光 编带机' + bs + '常规生产分光编带机图',
        '910编带机': '分光 编带机' + bs + '常规生产分光编带机图',
    }
    SOURCE_PARENT = {
        # 1838 条用 源: 字段的文档（来自 早期分光编带机图\2835分光机\...）
        '2835分光机': '分光 编带机' + bs + '早期分光编带机图',
    }
    # 机械图8大类的所有类别
    MACHINE_CATEGORIES = [
        '上下板机', '分光 编带机', '固晶机', '在线式打码机',
        '搅拌机', '支架贴胶带机', '点胶机', '贴片机',
    ]

    def wrap_folder_path(folder_path):
        """根据 FOLDER_PARENT 自动给顶层叶节点加上父目录"""
        if not folder_path or folder_path == '(根目录)':
            return folder_path
        top = folder_path.split(bs)[0]
        if top in FOLDER_PARENT:
            return FOLDER_PARENT[top] + bs + folder_path
        # 如果已含父目录（如 分光 编带机\710分光机），原样
        if top in ('分光 编带机', '上下板机', '固晶机', '在线式打码机', '搅拌机', '支架贴胶带机', '点胶机', '贴片机'):
            return folder_path
        return folder_path

    tree = defaultdict(list)  # path → [doc]
    folder_to_children = defaultdict(set)  # parent_folder → set(child_folder)
    folder_doc_count = defaultdict(int)
    docs_in_selected = []

    for d in Document.query.all():
        desc = d.description or ''
        # 支持两种格式：路径: 和 源:
        m = re.search(r'(?:路径|源):\s*(.+?)(\s*\||$)', desc)
        if not m:
            continue
        path = m.group(1).strip()
        if not path:
            continue
        # 计算文件夹路径（去掉文件名）
        segs = path.split(bs)
        folder_path = bs.join(segs[:-1])  # 不含文件名
        if not folder_path:
            folder_path = '(根目录)'

        # 应用父目录包装
        if '源:' in desc:
            # 1838 条文档：第1段可能是 8 大类之一（点胶机/固晶机等），或 2835分光机等
            top = folder_path.split(bs)[0]
            if top in SOURCE_PARENT:
                folder_path = SOURCE_PARENT[top] + bs + folder_path
            elif top in MACHINE_CATEGORIES:
                pass  # 已经是 8 大类之一，保持原样
            else:
                # 未知顶层 → 归到"分光 编带机 > 早期分光编带机图"下
                folder_path = '分光 编带机' + bs + '早期分光编带机图' + bs + folder_path
        else:
            folder_path = wrap_folder_path(folder_path)

        tree[folder_path].append(d)
        folder_doc_count[folder_path] += 1

        # 记录父-子关系
        parts = folder_path.split(bs)
        for i in range(len(parts)):
            parent = bs.join(parts[:i]) if i > 0 else ''
            child = parts[i]
            folder_to_children[parent].add(child)

    # 2) 如果指定了 path → 列出该路径下文档
    if selected_path:
        # 包含自身及所有子文件夹的文档
        if selected_path in tree:
            docs_in_selected = list(tree[selected_path])
        else:
            docs_in_selected = []
        # 加上所有子文件夹的文档
        for fp, items in tree.items():
            if fp.startswith(selected_path + bs):
                docs_in_selected.extend(items)

        # 关键词过滤
        if keyword:
            kw = keyword.lower()
            docs_in_selected = [d for d in docs_in_selected
                                if kw in (d.title or '').lower()
                                or kw in (d.name or '').lower()
                                or kw in (d.file_name or '').lower()]

        # 分类过滤
        if category_id:
            docs_in_selected = [d for d in docs_in_selected if d.category_id == category_id]

    # 3) 构造树节点（用于显示）
    # 路径 → 子文件夹列表（按字母排序）
    nodes = []  # [{name, path, depth, doc_count, has_children}]
    all_folders = sorted(set(tree.keys()))
    # 用前缀索引加速
    def collect_subfolders(parent_path):
        children = []
        prefix = parent_path + bs if parent_path and parent_path != '(根目录)' else ''
        for fp in all_folders:
            if parent_path == '(根目录)' and bs in fp:
                top = fp.split(bs)[0]
                if top not in children:
                    children.append(top)
            elif parent_path and fp.startswith(prefix):
                rest = fp[len(prefix):]
                if rest:
                    # 只取第一级子文件夹（多级时只保留首个 \\ 之前）
                    first_child = rest.split(bs)[0]
                    if first_child and first_child not in children:
                        children.append(first_child)
            elif not parent_path and bs not in fp and fp not in children:
                children.append(fp)
        return sorted(children)

    if not selected_path:
        # 根视图：显示所有顶层文件夹（含机械图8大类的空父目录）
        top = collect_subfolders('')
        for cat in MACHINE_CATEGORIES:
            if cat not in top:
                top.append(cat)
        top = sorted(set(top))
        for name in top:
            full_path = name  # 顶层文件夹
            count = folder_doc_count.get(full_path, 0)
            # 加上下属的子文件夹
            for fp in all_folders:
                if fp.startswith(name + bs):
                    count += folder_doc_count.get(fp, 0)
            nodes.append({
                'name': name,
                'path': full_path,
                'depth': 0,
                'doc_count': count,
            })
    else:
        # 子视图：显示当前路径下的子文件夹 + 文档
        sub = collect_subfolders(selected_path)
        for name in sub:
            child_path = (selected_path + bs + name) if selected_path and selected_path != '(根目录)' else name
            count = folder_doc_count.get(child_path, 0)
            for fp in all_folders:
                if fp.startswith(child_path + bs):
                    count += folder_doc_count.get(fp, 0)
            nodes.append({
                'name': name,
                'path': child_path,
                'depth': selected_path.count(bs) + 1,
                'doc_count': count,
            })

    # 4) 面包屑
    if selected_path:
        parts = selected_path.split(bs) if selected_path != '(根目录)' else []
        breadcrumbs = []
        for i, p in enumerate(parts):
            path_so_far = bs.join(parts[:i+1])
            breadcrumbs.append({'name': p, 'path': path_so_far})
    else:
        breadcrumbs = []

    # 5) 类别供筛选
    categories = DocumentCategory.query.all()

    # 保存当前完整树形浏览 URL 到 session，供详情页"返回列表"使用
    if selected_path:
        session['docs_tree_url'] = url_for('documents.tree_view', path=selected_path, q=keyword, category_id=category_id)
    else:
        session['docs_tree_url'] = url_for('documents.tree_view', q=keyword, category_id=category_id)
    # 标记最近一次访问是树形
    session['docs_last_view'] = 'tree'

    return render_template('documents/tree.html',
                           nodes=nodes,
                           docs=docs_in_selected,
                           selected_path=selected_path,
                           selected_path_display=(selected_path or '').replace(chr(92), ' / '),
                           breadcrumbs=breadcrumbs,
                           total_count=len(docs_in_selected),
                           categories=categories,
                           keyword=keyword,
                           category_id=category_id)


@document_bp.route('/<int:id>')
@login_required
def view_doc(id):
    doc = Document.query.get_or_404(id)
    # 根据最近一次访问决定返回：树形或列表
    if session.get('docs_last_view') == 'tree' and session.get('docs_tree_url'):
        back_url = session.get('docs_tree_url')
    else:
        back_url = session.get('docs_list_url') or url_for('documents.list_docs')
    return render_template('documents/view.html', doc=doc, back_url=back_url)


@document_bp.route('/<int:id>/download')
@login_required
def download_doc(id):
    """下载文档源文件（带原文件名）"""
    from flask import current_app, send_from_directory, abort
    from urllib.parse import quote
    doc = Document.query.get_or_404(id)
    if not doc.file_path or not doc.file_name:
        abort(404, '该文档未上传文件')
    # file_path 通常是相对路径如 "pdf/xxx.pdf"，UPLOAD_FOLDER 是绝对根
    upload_root = current_app.config.get('UPLOAD_FOLDER', '')
    return send_from_directory(
        upload_root,
        doc.file_path,
        as_attachment=True,
        download_name=doc.file_name
    )


@document_bp.route('/<int:id>/version/<int:version_id>/download')
@login_required
def download_version(id, version_id):
    """下载文档版本文件"""
    from flask import current_app, send_from_directory, abort
    doc = Document.query.get_or_404(id)
    ver = DocumentVersion.query.filter_by(id=version_id, document_id=doc.id).first_or_404()
    if not ver.file_path or not ver.file_name:
        abort(404, '该版本未上传文件')
    upload_root = current_app.config.get('UPLOAD_FOLDER', '')
    return send_from_directory(
        upload_root,
        ver.file_path,
        as_attachment=True,
        download_name=ver.file_name
    )


@document_bp.route('/<int:id>/edit', methods=['GET', 'POST'])
@login_required
def edit_doc(id):
    doc = Document.query.get_or_404(id)
    if request.method == 'POST':
        drawing_no = request.form.get('title', '').strip()
        name = request.form.get('name', '').strip()
        # 校验图号：8 位新格式 或 9 位旧格式（兼容老数据）
        import re
        if not (re.match(r'^\d{2}-\d{2}-\d{2}-\d{2}$', drawing_no) or
                re.match(r'^\d{3}-\d{3}-\d{3}$', drawing_no) or
                not drawing_no):
            flash('图号格式错误：8 位 XX-XX-XX-XX 或 9 位 XXX-XXX-XXX', 'danger')
        elif drawing_no and drawing_no in name:
            flash(f'名称中不允许包含图号「{drawing_no}」', 'danger')
        else:
            doc.title = drawing_no
            doc.name = name
            doc.description = request.form.get('description', '')
            doc.category_id = request.form.get('category_id', type=int)
            doc.tags = request.form.get('tags', '')
            file = request.files.get('file')
            if file and file.filename:
                fname, orig_name, fsize = save_file(file)
                if fname:
                    doc.file_name = orig_name
                    doc.file_path = fname
                    doc.file_size = fsize
            db.session.commit()
            # 注：不再自动同步 Product.code（避免覆盖原图号），仅做外键绑定
            flash('文档更新成功', 'success')
            return redirect(url_for('documents.view_doc', id=id))
    categories = DocumentCategory.query.all()
    return render_template('documents/edit.html', doc=doc, categories=categories)


@document_bp.route('/<int:id>/submit-approval', methods=['POST'])
@login_required
def submit_approval(id):
    doc = Document.query.get_or_404(id)
    # viewer 不可提交审批
    if current_user.role == 'viewer':
        flash('查看者无权提交审批', 'danger')
        return redirect(url_for('documents.view_doc', id=id))
    # 防止重复提交审批
    if doc.status in ('review', 'approved', 'published'):
        flash('该文档已在审批流程中或已审批完成', 'warning')
        return redirect(url_for('documents.view_doc', id=id))
    # 关键修复：先清掉该文档的所有历史审批记录（防止驳回重提时出现重复记录）
    DocApproval.query.filter_by(document_id=id).delete()
    doc.status = 'review'
    # 创建审批步骤：每个经理审批一次，没有经理时指定管理员
    managers = User.query.filter_by(role='manager').all()
    if not managers:
        managers = User.query.filter_by(role='admin').all()
    if not managers:
        flash('系统中没有经理或管理员角色用户，无法提交审批', 'danger')
        return redirect(url_for('documents.view_doc', id=id))
    for i, m in enumerate(managers):
        appr = DocApproval(document_id=id, step=i + 1, approver_id=m.id)
        db.session.add(appr)
    db.session.commit()
    flash('已提交审批', 'success')
    return redirect(url_for('documents.view_doc', id=id))


@document_bp.route('/<int:id>/delete', methods=['POST'])
@login_required
def delete_doc(id):
    doc = Document.query.get_or_404(id)
    # 只有作者本人或管理员可删除
    if doc.author_id != current_user.id and current_user.role != 'admin':
        flash('只有创建者或管理员才能删除文档', 'danger')
        return redirect(url_for('documents.view_doc', id=id))
    # 已发布文档需特殊处理
    if doc.status == 'published':
        flash('已发布的文档不能直接删除，请先作废', 'warning')
        return redirect(url_for('documents.view_doc', id=id))
    # 删除关联文件
    if doc.file_path:
        try:
            import os as _os
            fpath = _os.path.join(current_app.config['UPLOAD_FOLDER'], doc.file_path)
            if _os.path.exists(fpath):
                _os.remove(fpath)
        except Exception:
            pass
    # 删除关联的版本和审批记录
    DocVersion.query.filter_by(document_id=id).delete()
    DocApproval.query.filter_by(document_id=id).delete()
    db.session.delete(doc)
    db.session.commit()
    flash(f'文档「{doc.title}」已删除', 'success')
    return redirect(url_for('documents.list_docs'))


@document_bp.route('/<int:id>/approve', methods=['POST'])
@login_required
def approve_doc(id):
    approval = DocApproval.query.filter_by(document_id=id, approver_id=current_user.id, status='pending').first()
    if approval:
        action = request.form.get('action', '')
        if action not in ('approved', 'rejected'):
            flash('无效的审批操作', 'danger')
            return redirect(url_for('documents.view_doc', id=id))
        approval.status = action
        approval.updated_at = datetime.now()
        approval.comment = request.form.get('comment', '')
        if action == 'approved':
            next_pending = DocApproval.query.filter_by(document_id=id, status='pending').first()
            if not next_pending:
                doc = Document.query.get(id)
                doc.status = 'approved'
        else:
            doc = Document.query.get(id)
            doc.status = 'draft'
        db.session.commit()
        flash(f'审批完成：{"通过" if action == "approved" else "驳回"}', 'success')
    return redirect(url_for('documents.view_doc', id=id))


@document_bp.route('/<int:id>/publish', methods=['POST'])
@login_required
def publish_doc(id):
    require_role('admin', 'manager')
    doc = Document.query.get_or_404(id)
    doc.status = 'published'
    db.session.commit()
    flash('文档已发布', 'success')
    return redirect(url_for('documents.view_doc', id=id))


@document_bp.route('/<int:id>/new-version', methods=['GET', 'POST'])
@login_required
def new_version(id):
    doc = Document.query.get_or_404(id)
    if request.method == 'POST':
        old_version = doc.version
        # Save current as version history
        dv = DocVersion(
            document_id=id, version=old_version,
            file_name=doc.file_name, file_path=doc.file_path,
            file_size=doc.file_size,
            change_note=request.form.get('change_note', ''),
            created_by=current_user.id
        )
        db.session.add(dv)
        # Create new version
        doc.new_version()
        file = request.files.get('file')
        if file and file.filename:
            fname, orig_name, fsize = save_file(file)
            if fname:
                doc.file_name = orig_name
                doc.file_path = fname
                doc.file_size = fsize
        db.session.commit()
        flash('新版本已创建', 'success')
        return redirect(url_for('documents.view_doc', id=id))
    return render_template('documents/new_version.html', doc=doc)


@document_bp.route('/<int:id>/lock', methods=['POST'])
@login_required
def lock_doc(id):
    doc = Document.query.get_or_404(id)
    if doc.is_locked and doc.locked_by != current_user.id:
        flash('文档已被其他人锁定', 'warning')
    else:
        doc.is_locked = True
        doc.locked_by = current_user.id
        db.session.commit()
        flash('文档已锁定', 'success')
    return redirect(url_for('documents.view_doc', id=id))


@document_bp.route('/<int:id>/unlock', methods=['POST'])
@login_required
def unlock_doc(id):
    doc = Document.query.get_or_404(id)
    if doc.locked_by == current_user.id or current_user.role == 'admin':
        doc.is_locked = False
        doc.locked_by = None
        db.session.commit()
        flash('文档已解锁', 'success')
    else:
        flash('只有锁定者或管理员可以解锁', 'warning')
    return redirect(url_for('documents.view_doc', id=id))


# ─── 文档分类管理 ───
@document_bp.route('/categories')
@login_required
def list_categories():
    cats = DocumentCategory.query.all()
    return render_template('documents/categories.html', cats=cats)


@document_bp.route('/categories/create', methods=['POST'])
@login_required
def create_category():
    require_role('admin', 'manager')
    cat = DocumentCategory(
        name=request.form['name'],
        parent_id=request.form.get('parent_id', type=int),
        description=request.form.get('description', '')
    )
    db.session.add(cat)
    db.session.commit()
    flash('分类创建成功', 'success')
    return redirect(url_for('documents.list_categories'))


# ════════════════════════════════
#  2. 流程管理
# ════════════════════════════════
@workflow_bp.route('/')
@login_required
def list_workflows():
    workflows = Workflow.query.order_by(Workflow.created_at.desc()).all()
    return render_template('workflows/list.html', workflows=workflows)


@workflow_bp.route('/<int:id>')
@login_required
def view_workflow(id):
    wf = Workflow.query.get_or_404(id)
    return render_template('workflows/view.html', workflow=wf)


@workflow_bp.route('/<int:id>/edit', methods=['GET', 'POST'])
@login_required
def edit_workflow(id):
    require_role('admin', 'manager')
    wf = Workflow.query.get_or_404(id)
    if request.method == 'POST':
        wf.name = request.form['name']
        wf.description = request.form.get('description', '')
        wf.model_type = request.form.get('model_type', 'document')
        wf.is_active = request.form.get('is_active') == 'on'
        # 重置步骤
        WorkflowStep.query.filter_by(workflow_id=wf.id).delete()
        step_names = request.form.getlist('step_name[]')
        step_roles = request.form.getlist('step_role[]')
        for i, (sn, sr) in enumerate(zip(step_names, step_roles)):
            if sn.strip():
                step = WorkflowStep(workflow_id=wf.id, seq=i + 1, name=sn, approver_role=sr)
                db.session.add(step)
        db.session.commit()
        flash('流程已更新', 'success')
        return redirect(url_for('workflows.list_workflows'))
    return render_template('workflows/create.html', workflow=wf, edit_mode=True)


@workflow_bp.route('/<int:id>/delete', methods=['POST'])
@login_required
def delete_workflow(id):
    require_role('admin', 'manager')
    wf = Workflow.query.get_or_404(id)
    WorkflowStep.query.filter_by(workflow_id=wf.id).delete()
    db.session.delete(wf)
    db.session.commit()
    flash(f'流程「{wf.name}」已删除', 'success')
    return redirect(url_for('workflows.list_workflows'))


@workflow_bp.route('/<int:id>/toggle', methods=['POST'])
@login_required
def toggle_workflow(id):
    require_role('admin', 'manager')
    wf = Workflow.query.get_or_404(id)
    wf.is_active = not wf.is_active
    db.session.commit()
    flash(f'流程「{wf.name}」已{"启用" if wf.is_active else "禁用"}', 'success')
    return redirect(url_for('workflows.list_workflows'))


@workflow_bp.route('/create', methods=['GET', 'POST'])
@login_required
def create_workflow():
    require_role('admin', 'manager')
    if request.method == 'POST':
        wf = Workflow(
            name=request.form['name'],
            description=request.form.get('description', ''),
            model_type=request.form.get('model_type', 'document')
        )
        db.session.add(wf)
        db.session.flush()
        step_names = request.form.getlist('step_name[]')
        step_roles = request.form.getlist('step_role[]')
        for i, (sn, sr) in enumerate(zip(step_names, step_roles)):
            if sn.strip():
                step = WorkflowStep(workflow_id=wf.id, seq=i + 1, name=sn, approver_role=sr)
                db.session.add(step)
        db.session.commit()
        flash('流程创建成功', 'success')
        return redirect(url_for('workflows.list_workflows'))
    return render_template('workflows/create.html')


# ════════════════════════════════
#  3. 结构管理
# ════════════════════════════════
@structure_bp.route('/')
@login_required
def tree():
    keyword = request.args.get('q', '').strip()
    level_filter = request.args.get('level', type=int)
    query = Product.query
    if keyword:
        like = f'%{keyword}%'
        query = query.filter(db.or_(
            Product.code.ilike(like),
            Product.name.ilike(like),
            Product.description.ilike(like)
        ))
    if level_filter is not None:
        query = query.filter_by(level=level_filter)
    # 默认显示根产品
    products = query.filter_by(parent_id=None).all()
    # 统计各级产品数量
    level_counts = {
        0: Product.query.filter_by(level=0).count(),
        1: Product.query.filter_by(level=1).count(),
        2: Product.query.filter_by(level=2).count(),
        3: Product.query.filter_by(level=3).count(),
    }
    return render_template('structures/tree.html', products=products,
                           keyword=keyword, level_filter=level_filter,
                           level_counts=level_counts)


@structure_bp.route('/create', methods=['GET', 'POST'])
@login_required
def create_product():
    require_role('admin', 'manager', 'user')  # viewer 不可创建产品
    if request.method == 'POST':
        # 系统主物料号（唯一标识）
        new_sid = gen_system_id()
        while Product.query.filter_by(system_id=new_sid).first():
            new_sid = gen_system_id()
        p = Product(
            system_id=new_sid,
            code=new_sid,  # code 跟随 system_id，不再单独维护原图号
            name=request.form['name'],
            description=request.form.get('description', ''),
            item_type=request.form.get('item_type', 'PART'),
            revision=request.form.get('revision', 'R00'),
            parent_id=request.form.get('parent_id', type=int),
            level=int(request.form.get('level', 0)),
            applicable_models=request.form.get('applicable_models', '')
        )
        db.session.add(p)
        db.session.commit()
        flash(f'产品创建成功！主物料号：{p.system_id}', 'success')
        return redirect(url_for('structures.product_detail', id=p.id))
    parents = Product.query.all()
    return render_template('structures/create.html', parents=parents)


@structure_bp.route('/<int:id>/detail')
@login_required
def product_detail(id):
    """产品详情页：显示基本信息、关联 BOM、关联文档、父子产品"""
    p = Product.query.get_or_404(id)
    # 关联的 eBOM
    eboms = Bom.query.filter_by(product_id=id, bom_type='EBOM').all()
    # 关联的 mBOM
    mboms = Bom.query.filter_by(product_id=id, bom_type='MBOM').all()
    # 父产品
    parent = Product.query.get(p.parent_id) if p.parent_id else None
    # 子产品
    children = p.children.all()
    # 反向引用：哪些 BOM 用此产品作为物料
    used_in_boms = db.session.query(Bom).join(BomItem).filter(BomItem.product_id == id).all()
    # 关联文档 — 优先外键绑定，回退为图号/文件名模糊匹配
    docs = []
    if p.document_id:
        doc = Document.query.get(p.document_id)
        if doc:
            docs = [doc]
    if not docs:
        identifiers = [p.system_id or '', p.code or '']
        identifier_list = [x for x in identifiers if x]
        seen_ids = set()
        for ident in identifier_list:
            rows = Document.query.filter(
                (Document.title.like(ident + '%')) | (Document.title.like(ident)) |
                (Document.file_name.like(ident + '%'))
            ).limit(20).all()
            for r in rows:
                if r.id not in seen_ids:
                    seen_ids.add(r.id)
                    docs.append(r)
        docs = docs[:20]
    # Odoo 对应：查询已同步的 mBOM（按 system_id 找产品 → 找 mBOM → 看 sync_status）
    odoo_synced_boms = []
    for mb in mboms:
        if mb.sync_status == 'synced':
            odoo_synced_boms.append(mb)
    return render_template('structures/detail.html', product=p,
                           eboms=eboms, mboms=mboms, parent=parent, children=children,
                           used_in_boms=used_in_boms, docs=docs,
                           odoo_synced_boms=odoo_synced_boms)


@structure_bp.route('/<int:id>/children')
@login_required
def get_children(id):
    p = Product.query.get_or_404(id)
    children = [{'id': c.id, 'code': c.code, 'system_id': c.system_id or '',
                 'document_id': c.document_id, 'name': c.name, 'level': c.level,
                 'has_children': c.children.count() > 0} for c in p.children]
    return jsonify(children)


@structure_bp.route('/<int:id>/delete', methods=['POST'])
@login_required
def delete_product(id):
    require_role('admin', 'manager')
    p = Product.query.get_or_404(id)
    if p.children.count() > 0:
        flash(f'「{p.name}」下还有 {p.children.count()} 个子产品，请先删除子产品', 'warning')
        return redirect(url_for('structures.tree'))
    # 检查是否被BOM引用
    if p.boms.count() > 0:
        flash(f'「{p.name}」已被 BOM 引用，不能删除', 'warning')
        return redirect(url_for('structures.tree'))
    db.session.delete(p)
    db.session.commit()
    flash(f'产品「{p.name}」已删除', 'success')
    return redirect(url_for('structures.tree'))


# ════════════════════════════════
#  4. BOM 管理 — eBOM 为源，mBOM 由 eBOM 结构转换生成
# ════════════════════════════════
@bom_bp.route('/')
@login_required
def list_boms():
    bom_type = request.args.get('type', '')  # EBOM / MBOM / all
    sync_filter = request.args.get('sync', '')  # not_synced / synced
    page = max(1, int(request.args.get('page', 1)))
    per_page = 20

    # 保存当前完整列表 URL 到 session，供详情页"返回列表"使用
    from urllib.parse import urlencode
    query_args = {}
    if bom_type: query_args['type'] = bom_type
    if sync_filter: query_args['sync'] = sync_filter
    if page and page != 1: query_args['page'] = page
    if query_args:
        session['boms_list_url'] = url_for('boms.list_boms') + '?' + urlencode(query_args)
    else:
        session['boms_list_url'] = url_for('boms.list_boms')

    query = Bom.query
    if bom_type:
        query = query.filter_by(bom_type=bom_type)
    if sync_filter:
        query = query.filter_by(sync_status=sync_filter)
    total = query.count()
    boms = query.order_by(Bom.updated_at.desc()).offset((page - 1) * per_page).limit(per_page).all()
    total_pages = max(1, (total + per_page - 1) // per_page)
    # 统计数据
    ebom_count = Bom.query.filter_by(bom_type='EBOM').count()
    mbom_count = Bom.query.filter_by(bom_type='MBOM').count()
    pending_sync = Bom.query.filter_by(bom_type='MBOM', sync_status='not_synced', status='released').count()
    my_pending_bom_count = BomApproval.query.filter_by(
        approver_id=current_user.id, status='pending').count()

    # 计算每个 BOM 当前用户是否需要审批
    pending_by_bom = {}
    for a in BomApproval.query.filter(
        BomApproval.approver_id == current_user.id, BomApproval.status == 'pending'
    ).all():
        pending_by_bom[a.bom_id] = a.step

    return render_template('boms/list.html', boms=boms,
                           bom_type=bom_type, sync_filter=sync_filter,
                           page=page, per_page=per_page, total=total, total_pages=total_pages,
                           ebom_count=ebom_count, mbom_count=mbom_count,
                           pending_sync=pending_sync,
                           my_pending_bom_count=my_pending_bom_count,
                           pending_by_bom=pending_by_bom)


@bom_bp.route('/create', methods=['GET', 'POST'])
@login_required
def create_bom():
    """创建 BOM — 默认为 eBOM（设计BOM），是产品数据的唯一源头"""
    require_role('admin', 'manager', 'user')  # viewer 不可创建
    if request.method == 'POST':
        bom_type = request.form.get('bom_type', 'EBOM')
        bom = Bom(
            name=request.form['name'],
            description=request.form.get('description', ''),
            bom_type=bom_type,
            product_id=request.form.get('product_id', type=int),
            created_by=current_user.id
        )
        db.session.add(bom)
        db.session.flush()
        prod_ids = request.form.getlist('product_id[]')
        qtys = request.form.getlist('qty[]')
        units = request.form.getlist('unit[]')
        notes = request.form.getlist('note[]')
        codes = request.form.getlist('code[]')
        part_types = request.form.getlist('part_type[]')
        for i, (pid, qty, unit) in enumerate(zip(prod_ids, qtys, units)):
            if pid and qty:
                note = notes[i] if i < len(notes) else ''
                code = codes[i] if i < len(codes) else ''
                ptype = part_types[i] if i < len(part_types) else ''
                bi = BomItem(bom_id=bom.id, product_id=int(pid),
                             quantity=float(qty), unit=unit, seq=i + 1, note=note,
                             code=code, part_type=ptype)
                db.session.add(bi)
        db.session.commit()
        type_label = 'eBOM（工程 BOM）' if bom_type == 'EBOM' else 'mBOM（制造 BOM）'
        flash(f'{type_label} 创建成功', 'success')
        return redirect(url_for('boms.list_boms'))
    products = Product.query.all()
    return render_template('boms/create.html', products=products)


# ── Excel 导入 eBOM ──
@bom_bp.route('/import-excel', methods=['POST'])
@login_required
def import_excel_bom():
    """上传 Excel 文件，解析其中物料自动创建 eBOM 和关联产品"""
    require_role('admin', 'manager', 'user')  # viewer 不可导入
    file = request.files.get('excel_file')
    if not file or not file.filename:
        flash('请选择一个 Excel 文件', 'warning')
        return redirect(url_for('boms.create_bom'))

    ext = file.filename.rsplit('.', 1)[-1].lower() if '.' in file.filename else ''
    if ext not in ('xls', 'xlsx'):
        flash('仅支持 .xls 或 .xlsx 格式', 'warning')
        return redirect(url_for('boms.create_bom'))

    import tempfile, os
    tmp_path = os.path.join(tempfile.gettempdir(), f'bom_import_{os.urandom(4).hex()}.{ext}')
    file.save(tmp_path)

    try:
        # 解析 Excel
        items = _parse_excel_bom(tmp_path)
        if not items:
            flash('Excel 中没有识别到有效的物料数据。请确保表格有"序号""名称""规格型号""用量"列。', 'warning')
            return redirect(url_for('boms.create_bom'))

        # 创建 eBOM
        bom_name = request.form.get('name', file.filename.rsplit('.', 1)[0])
        bom_type = request.form.get('bom_type', 'EBOM')
        product_id = request.form.get('product_id', type=int)

        bom = Bom(
            name=bom_name,
            description=f'从 Excel 导入：{file.filename}',
            bom_type=bom_type,
            product_id=product_id,
            created_by=current_user.id,
            status='draft'
        )
        db.session.add(bom)
        db.session.flush()

        # 合并同一规格型号的物料行
        merged = {}
        for item in items:
            key = item['spec'] if item['spec'] else item['name']
            if key in merged:
                merged[key]['usage'] += item['usage']
            else:
                merged[key] = item

        item_count = 0
        new_product_count = 0
        for seq, (key, item) in enumerate(merged.items(), 1):
            spec = item['spec'] if item['spec'] else item['name']
            # 新建或复用产品
            existing = Product.query.filter_by(code=spec).first()
            if not existing:
                existing = Product(
                    code=spec[:80], name=item['name'][:200],
                    level=3, status='active'
                )
                db.session.add(existing)
                db.session.flush()
                new_product_count += 1

            bi = BomItem(
                bom_id=bom.id, product_id=existing.id,
                quantity=item['usage'], unit=item['unit'],
                seq=seq,
                note=f"{item['brand']} | {item['supplier']} | {item['note']}"[:200]
            )
            db.session.add(bi)
            item_count += 1

        db.session.commit()
        flash(f'导入成功！创建 eBOM「{bom_name}」共 {item_count} 项物料，自动创建 {new_product_count} 个新产品。', 'success')
        return redirect(url_for('boms.view_bom', id=bom.id))

    except Exception as e:
        import logging
        logging.getLogger(__name__).error(f'Excel import failed: {e}', exc_info=True)
        flash(f'导入失败：{str(e)[:100]}', 'danger')
        return redirect(url_for('boms.create_bom'))
    finally:
        try:
            os.remove(tmp_path)
        except:
            pass


def _parse_excel_bom(filepath):
    """解析 Excel BOM，返回物料列表（同 import_mechanical.py 的 parse_excel_bom）"""
    ext = filepath.rsplit('.', 1)[-1].lower() if '.' in filepath else ''

    if ext == 'xls':
        import xlrd
        wb = xlrd.open_workbook(filepath)
        sheets = [(sn, wb.sheet_by_name(sn)) for sn in wb.sheet_names()]
    else:
        import openpyxl
        wb = openpyxl.load_workbook(filepath, data_only=True)
        sheets = [(sn, wb[sn]) for sn in wb.sheetnames]

    items = []
    for sn, sh in sheets:
        if ext == 'xls':
            nrows, ncols = sh.nrows, sh.ncols
            cell = lambda r, c: str(sh.cell_value(r, c)).strip()
        else:
            nrows, ncols = sh.max_row + 1, sh.max_column + 1
            cell = lambda r, c: str(sh.cell(r, c).value or '').strip()

        # 找表头行
        header_row = None
        header = []
        for r in range(min(nrows, 30)):
            row_vals = [cell(r, c) for c in range(min(ncols, 16))]
            if any('序号' in v for v in row_vals):
                header_row = r
                header = row_vals
                break
        if header_row is None:
            continue

        col_map = {}
        for i, h in enumerate(header):
            if '序号' in h: col_map['seq'] = i
            elif '名称' in h: col_map['name'] = i
            elif '规格型号' in h or '型号' in h: col_map['spec'] = i
            elif '品牌' in h: col_map['brand'] = i
            elif '用量' in h: col_map['usage'] = i
            elif '数量' in h and '用量' not in h: col_map['qty'] = i
            elif '单位' in h: col_map['unit'] = i
            elif '供应商' in h: col_map['supplier'] = i
            elif '备注' in h: col_map['note'] = i

        if 'name' not in col_map and 'spec' not in col_map:
            continue

        for r in range(header_row + 1, nrows):
            seq_val = cell(r, col_map.get('seq', 0))
            name_val = cell(r, col_map.get('name', 1))
            spec_val = cell(r, col_map.get('spec', 2))

            if not seq_val or seq_val == '0.0':
                continue

            try:
                float(seq_val)
            except ValueError:
                continue

            if not name_val and not spec_val:
                continue

            usage = 1.0
            if 'usage' in col_map:
                try:
                    usage = float(cell(r, col_map['usage']))
                except (ValueError, TypeError):
                    usage = 1.0
            elif 'qty' in col_map:
                try:
                    usage = float(cell(r, col_map['qty']))
                except (ValueError, TypeError):
                    usage = 1.0

            unit = cell(r, col_map['unit']) if 'unit' in col_map else '个'
            if not unit or unit == '':
                unit = '个'

            items.append({
                'name': name_val,
                'spec': spec_val,
                'brand': cell(r, col_map['brand']) if 'brand' in col_map else '',
                'usage': usage,
                'unit': unit,
                'supplier': cell(r, col_map['supplier']) if 'supplier' in col_map else '',
                'note': cell(r, col_map['note']) if 'note' in col_map else '',
            })

    return items


@bom_bp.route('/<int:id>')
@login_required
def view_bom(id):
    bom = Bom.query.get_or_404(id)
    # 从 session 读取列表页 URL（含筛选状态）
    back_url = session.get('boms_list_url') or url_for('boms.list_boms')
    products = Product.query.all()
    # 获取该 eBOM 派生的 mBOM 列表
    derived_mboms = []
    # 关联文档
    linked_docs = bom.linked_docs.all() if hasattr(bom, 'linked_docs') else []
    # 审批记录
    approvals = bom.approvals.all() if hasattr(bom, 'approvals') else []
    # 同系列版本（用于版本对比）
    sibling_boms = Bom.query.filter(
        Bom.id != bom.id,
        Bom.product_id == bom.product_id,
        Bom.bom_type == bom.bom_type
    ).order_by(Bom.version.desc()).limit(20).all()
    # 全局文档/产品列表（Tab 选择用）
    all_docs = Document.query.order_by(Document.title).limit(200).all()
    all_products = Product.query.order_by(Product.name).limit(500).all()

    # 按 BOM 上下文过滤候选文档（用于文档 Tab 下拉框）
    candidate_docs = []
    seen_doc_ids = {ld.document_id for ld in linked_docs}
    import re
    keywords = set()
    if bom.product:
        keywords.add(bom.product.name.split(' ')[0])
        for m in re.findall(r'\d+', bom.product.name):
            if len(m) >= 2: keywords.add(m)
        if bom.product.code and bom.product.code not in ('系统未生成', ''):
            keywords.add(bom.product.code)
    for item in bom.items.all():
        if item.product and item.product.code:
            for n in (10, 8, 6):
                c = item.product.code[:n]
                if c and len(c) >= 4 and not (c.isdigit() and len(c) < 5):
                    keywords.add(c)
        if item.product:
            for m in re.findall(r'\d+', item.product.name):
                if len(m) >= 4: keywords.add(m)
    for kw in keywords:
        if not kw or len(kw) < 2:
            continue
        docs = Document.query.filter(Document.file_name.ilike(f'%{kw}%')).limit(30).all()
        for d in docs:
            if d.id not in seen_doc_ids and d not in candidate_docs:
                candidate_docs.append(d)
    candidate_docs = candidate_docs[:100]  # 最多 100 条，避免选项过长

    # 组件分页
    items_page = request.args.get('items_page', 1, type=int)
    items_per_page = 20
    items_total = bom.items.count()
    items_total_pages = max(1, (items_total + items_per_page - 1) // items_per_page)
    items = bom.items.order_by(BomItem.seq).offset((items_page - 1) * items_per_page).limit(items_per_page).all()

    # 关联文档分页
    docs_page = request.args.get('docs_page', 1, type=int)
    docs_per_page = 20
    docs_total = len(linked_docs)
    docs_total_pages = max(1, (docs_total + docs_per_page - 1) // docs_per_page)
    linked_docs_paged = linked_docs[(docs_page - 1) * docs_per_page : docs_page * docs_per_page]

    return render_template('boms/view.html', bom=bom, products=products,
                           derived_mboms=derived_mboms, linked_docs=linked_docs_paged,
                           approvals=approvals, sibling_boms=sibling_boms,
                           all_docs=all_docs, all_products=all_products,
                           candidate_docs=candidate_docs,
                           items=items, items_page=items_page, items_per_page=items_per_page,
                           items_total=items_total, items_total_pages=items_total_pages,
                           docs_page=docs_page, docs_per_page=docs_per_page,
                           docs_total=docs_total, docs_total_pages=docs_total_pages,
                           back_url=back_url)


@bom_bp.route('/<int:id>/edit', methods=['GET', 'POST'])
@login_required
def edit_bom(id):
    bom = Bom.query.get_or_404(id)
    require_bom_write(bom)
    if bom.status == 'released':
        flash('已发布的 BOM 不能编辑，请先作废或创建新版本', 'warning')
        return redirect(url_for('boms.view_bom', id=id))
    if request.method == 'POST':
        bom.name = request.form['name']
        bom.description = request.form.get('description', '')
        bom.bom_type = request.form.get('bom_type', 'EBOM')
        bom.product_id = request.form.get('product_id', type=int)
        # 重建物料明细
        BomItem.query.filter_by(bom_id=bom.id).delete()
        prod_ids = request.form.getlist('product_id[]')
        qtys = request.form.getlist('qty[]')
        units = request.form.getlist('unit[]')
        notes = request.form.getlist('note[]')
        codes = request.form.getlist('code[]')
        part_types = request.form.getlist('part_type[]')
        for i, (pid, qty, unit) in enumerate(zip(prod_ids, qtys, units)):
            if pid and qty:
                note = notes[i] if i < len(notes) else ''
                code = codes[i] if i < len(codes) else ''
                ptype = part_types[i] if i < len(part_types) else ''
                bi = BomItem(bom_id=bom.id, product_id=int(pid),
                             quantity=float(qty), unit=unit, seq=i + 1, note=note,
                             code=code, part_type=ptype)
                db.session.add(bi)
        db.session.commit()
        flash('BOM 已更新', 'success')
        return redirect(url_for('boms.view_bom', id=id))
    products = Product.query.all()
    return render_template('boms/create.html', bom=bom, products=products, edit_mode=True)


@bom_bp.route('/<int:id>/submit-approval', methods=['POST'])
@login_required
def bom_submit_approval(id):
    """提交 BOM 进入审批流程"""
    bom = Bom.query.get_or_404(id)
    require_bom_write(bom)
    if bom.status != 'draft':
        flash('只有草稿状态才能提交审批', 'warning')
        return redirect(url_for('boms.view_bom', id=id))
    bom.status = 'review'
    BomApproval.query.filter_by(bom_id=id).delete()
    managers = User.query.filter_by(role='manager').all()
    sub_id = bom.created_by or current_user.id
    for i, m in enumerate(managers):
        db.session.add(BomApproval(bom_id=id, step=i+1, approver_id=m.id, submitter_id=sub_id))
    if not managers:
        admin_user = User.query.filter_by(role='admin').first()
        db.session.add(BomApproval(bom_id=id, step=1, approver_id=admin_user.id if admin_user else current_user.id, submitter_id=sub_id))
    db.session.commit()
    flash(f'BOM 已提交审批，{len(managers) or 1} 位审批人已加入', 'success')
    return redirect(url_for('boms.view_bom', id=id))


@bom_bp.route('/<int:id>/release', methods=['POST'])
@login_required
def release_bom(id):
    require_role('admin', 'manager')
    bom = Bom.query.get_or_404(id)
    if bom.status == 'released':
        flash('该 BOM 已经发布', 'warning')
    elif bom.items.count() == 0:
        flash('BOM 没有物料明细，不能发布', 'warning')
    else:
        bom.status = 'released'
        # 如果是 eBOM，提示可以转换为 mBOM
        if bom.bom_type == 'EBOM':
            db.session.commit()
            flash(f'eBOM「{bom.name}」已发布。可点击"转为 mBOM"创建制造 BOM。', 'success')
        else:
            db.session.commit()
            flash(f'mBOM「{bom.name}」已发布。可推送到 Odoo 系统。', 'success')
        return redirect(url_for('boms.view_bom', id=id))
    return redirect(url_for('boms.view_bom', id=id))


@bom_bp.route('/<int:id>/obsolete', methods=['POST'])
@login_required
def obsolete_bom(id):
    require_role('admin', 'manager')
    bom = Bom.query.get_or_404(id)
    bom.status = 'obsolete'
    db.session.commit()
    flash(f'BOM「{bom.name}」已作废', 'success')
    return redirect(url_for('boms.view_bom', id=id))


@bom_bp.route('/<int:id>/delete', methods=['POST'])
@login_required
def delete_bom(id):
    bom = Bom.query.get_or_404(id)
    require_bom_write(bom)
    if bom.status == 'released' and current_user.role != 'admin':
        flash('已发布的 BOM 需管理员才能删除', 'warning')
        return redirect(url_for('boms.view_bom', id=id))
    # 检查是否有派生 mBOM
    if bom.bom_type == 'EBOM':
        derived = Bom.query.filter_by(source_ebom_id=id).all()
        if derived:
            flash(f'该 eBOM 已有 {len(derived)} 个派生 mBOM，请先删除 mBOM 再删除 eBOM', 'warning')
            return redirect(url_for('boms.view_bom', id=id))
    BomApproval.query.filter_by(bom_id=id).delete()
    BomItem.query.filter_by(bom_id=id).delete()
    BomDocument.query.filter_by(bom_id=id).delete()
    Bom.query.filter_by(source_ebom_id=id).update({'source_ebom_id': None})
    BomConversion.query.filter(db.or_(
        BomConversion.source_bom_id == id,
        BomConversion.target_bom_id == id
    )).delete()
    db.session.delete(bom)
    db.session.commit()
    flash(f'BOM「{bom.name}」已删除', 'success')
    return redirect(url_for('boms.list_boms'))


# ── 4a. eBOM → mBOM 结构转换 ──
@bom_bp.route('/<int:id>/convert-to-mbom', methods=['GET', 'POST'])
@login_required
def convert_to_mbom(id):
    """将 eBOM 结构变换为 mBOM（仅做结构变换，不含工序/工时/损耗率）"""
    require_role('admin', 'manager', 'user')  # viewer 不可转换
    ebom = Bom.query.get_or_404(id)
    if ebom.bom_type != 'EBOM':
        flash('只有 eBOM 才能转换为 mBOM', 'warning')
        return redirect(url_for('boms.view_bom', id=id))
    if ebom.status != 'released':
        flash('只有已发布的 eBOM 才能转换为 mBOM', 'warning')
        return redirect(url_for('boms.view_bom', id=id))
    if ebom.items.count() == 0:
        flash('该 eBOM 没有物料明细，无法转换为 mBOM', 'warning')
        return redirect(url_for('boms.view_bom', id=id))

    if request.method == 'POST':
        # 创建 mBOM，初始结构 = eBOM 结构
        mbom = Bom(
            name=request.form.get('name', ebom.name + ' - 制造BOM'),
            description=request.form.get('description', f'由 {ebom.bom_no} 结构变换而来'),
            bom_type='MBOM',
            product_id=ebom.product_id,
            source_ebom_id=ebom.id,
            created_by=current_user.id
        )
        db.session.add(mbom)
        db.session.flush()

        # 复制 eBOM 物料明细到 mBOM（可在此调整结构）
        item_count = 0
        for item in ebom.items:
            bi = BomItem(
                bom_id=mbom.id,
                product_id=item.product_id,
                quantity=item.quantity,
                unit=item.unit,
                reference=item.reference,
                seq=item.seq,
                note=item.note,
                code=item.code,
                part_type=item.part_type
            )
            db.session.add(bi)
            item_count += 1

        # 记录转换历史
        conv = BomConversion(
            source_bom_id=ebom.id,
            target_bom_id=mbom.id,
            converted_by=current_user.id,
            conversion_note=request.form.get('conversion_note', f'标准结构变换 — {item_count} 项物料')
        )
        db.session.add(conv)
        db.session.commit()
        flash(f'mBOM「{mbom.name}」已从 eBOM 创建成功，共 {item_count} 项物料。请进入详情页调整结构后发布。', 'success')
        return redirect(url_for('boms.view_bom', id=mbom.id))

    # GET: 展示转换预览页
    return render_template('boms/convert.html', ebom=ebom)

@bom_bp.route('/<int:id>/sync-status', methods=['GET'])
@login_required
def bom_sync_status(id):
    bom = Bom.query.get_or_404(id)
    return jsonify({
        'sync_status': bom.sync_status or 'not_synced',
        'sync_time': bom.sync_time.isoformat() if bom.sync_time else None,
        'sync_message': bom.sync_message or ''
    })





@bom_bp.route('/<int:id>/approve-step/<int:step>', methods=['POST'])
@login_required
def bom_approve_step(id, step):
    appr = BomApproval.query.filter_by(bom_id=id, step=step, status='pending').first_or_404()
    if appr.approver_id != current_user.id:
        flash('非当前审批人，无法审批', 'danger')
        return redirect(url_for('boms.view_bom', id=id))
    appr.status = 'approved'
    appr.decided_at = datetime.now()
    appr.comment = request.form.get('comment', '')
    remaining = BomApproval.query.filter_by(bom_id=id, status='pending').count()
    if remaining == 0:
        Bom.query.filter_by(id=id).update({'status': 'approved'})
    db.session.commit()
    flash(f'步骤 {step} 审批通过', 'success')
    return redirect(url_for('boms.view_bom', id=id))

@bom_bp.route('/<int:id>/reject-step/<int:step>', methods=['POST'])
@login_required
def bom_reject_step(id, step):
    appr = BomApproval.query.filter_by(bom_id=id, step=step, status='pending').first_or_404()
    if appr.approver_id != current_user.id:
        flash('非当前审批人，无法审批', 'danger')
        return redirect(url_for('boms.view_bom', id=id))
    appr.status = 'rejected'
    appr.decided_at = datetime.now()
    appr.comment = request.form.get('comment', '')
    Bom.query.filter_by(id=id).update({'status': 'rejected'})
    db.session.commit()
    flash(f'步骤 {step} 已驳回', 'danger')
    return redirect(url_for('boms.view_bom', id=id))


# ── 4b. mBOM 推送到 Odoo（纯单向，只推不取） ──
@bom_bp.route('/<int:id>/push-to-odoo', methods=['POST'])
@login_required
def push_to_odoo(id):
    """将 mBOM 单向推送到 Odoo（PLM → Odoo，不拉取数据）"""
    require_role('admin', 'manager')
    bom = Bom.query.get_or_404(id)
    if bom.bom_type != 'MBOM':
        flash('只有 mBOM 才能推送到 Odoo', 'warning')
        return redirect(url_for('boms.view_bom', id=id))
    if bom.status != 'released':
        flash('只有已发布的 mBOM 才能推送到 Odoo', 'warning')
        return redirect(url_for('boms.view_bom', id=id))

    # 查找活跃的 Odoo 配置
    odoo_config = IntegrationConfig.query.filter_by(
        system_type='odoo', is_active=True
    ).first()

    if not odoo_config:
        flash('未找到活跃的 Odoo 集成配置，请先在「系统集成」中添加并启用 Odoo 连接', 'warning')
        return redirect(url_for('boms.view_bom', id=id))

    # ── Odoo 推送通过独立子进程 push_to_odoo_runner.py 执行 ──
    import logging as _lg, subprocess, os, sys
    # 防重入：使用 DB 级别原子条件更新防止并发重复推送
    _bom_id = bom.id
    _cfg_id = odoo_config.id
    _base_dir = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
    _db_path = os.path.join(_base_dir, 'plm.db')
    if bom.sync_status == 'synced' and bom.sync_time and \
       (datetime.now() - bom.sync_time).total_seconds() < 300:
        flash('该 BOM 在 5 分钟内已同步成功，无需重复推送', 'info')
        return redirect(url_for('boms.view_bom', id=id))
    # 原子 CAS：仅当状态非 pushing 时才抢夺执行权
    affected = Bom.query.filter_by(id=_bom_id).filter(
        Bom.sync_status.in_(['not_synced', 'sync_failed', 'synced'])
    ).update({'sync_status': 'pushing', 'sync_message': 'dispatching', 'sync_time': datetime.now()},
             synchronize_session='fetch')
    db.session.commit()
    if affected == 0:
        flash('该 BOM 正在推送中或已被其他请求抢占，请稍后查看', 'warning')
        return redirect(url_for('boms.view_bom', id=id))
    runner = os.path.join(_base_dir, "push_to_odoo_runner.py")
    log_file = os.path.join(_base_dir, "plm_push.log")
    proc = subprocess.Popen([sys.executable, runner, str(_bom_id), str(_cfg_id), log_file, _db_path],
                            stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
    bom.sync_message = str(proc.pid)
    db.session.commit()
    _lg.getLogger().info(f"Push dispatched: bom={_bom_id} pid={proc.pid}")
    flash(f'<i class="bi bi-hourglass-split me-1"></i>推送任务已派发(PID={proc.pid})，1-2 分钟后刷新本页查看结果。', 'info')
    return redirect(url_for('boms.view_bom', id=id))




# ── API: 编码生成 ──
@bom_bp.route('/api/generate-code', methods=['POST'])
@login_required
def api_generate_code():
    """根据分类参数自动生成配件编码"""
    import json
    from app.code_generator import generate_part_code, generate_std_code, infer_family_code
    
    category = request.form.get('category', '01')       # 大类码
    family = request.form.get('family', '99')            # 产品族码
    component = request.form.get('component', '01')      # 部件分类码
    bom_name = request.form.get('bom_name', '')          # BOM名称（用于推断产品族）
    
    # 如果没传产品族码，从 BOM 名称推断
    if family == '99' and bom_name:
        family = infer_family_code(bom_name)
    
    code = generate_part_code(BomItem, family, component, category)
    return json.dumps({'code': code, 'family': family})

@bom_bp.route('/api/mappings')
@login_required
def api_mappings():
    """返回编码体系的所有映射表（供前端用）"""
    from app.code_generator import get_all_mappings
    return get_all_mappings()

@bom_bp.route('/<int:id>/set-item-code', methods=['POST'])
@login_required
def set_item_code(id):
    """为 BOM 中的物料行设置编码"""
    bom = Bom.query.get_or_404(id)
    item_id = request.form.get('item_id', type=int)
    code = request.form.get('code', '').strip()
    part_type = request.form.get('part_type', '').strip()
    
    item = BomItem.query.filter_by(id=item_id, bom_id=bom.id).first_or_404()
    item.code = code if code else None
    item.part_type = part_type if part_type else None
    db.session.commit()
    
    flash(f'编码 {code} 已保存', 'success')
    return redirect(url_for('boms.view_bom', id=id))

# ── 变更管理 ──
@change_bp.route('/')
@login_required
def list_changes():
    # viewer 只能看到自己提交的变更
    if current_user.role == 'viewer':
        changes = ChangeRequest.query.filter_by(applicant_id=current_user.id).order_by(ChangeRequest.created_at.desc()).limit(50).all()
    else:
        changes = ChangeRequest.query.order_by(ChangeRequest.created_at.desc()).limit(50).all()
    return render_template('changes/list.html', changes=changes)

# ── 项目管理 ──
@project_bp.route('/')
@login_required
def list_projects():
    projects = Project.query.order_by(Project.created_at.desc()).all()
    return render_template('projects/list.html', projects=projects)

# ── 工艺路线 ──
@process_bp.route('/')
@login_required
def list_processes():
    processes = []
    return render_template('processes/list.html', processes=processes)

# ── 系统集成 ──
@integration_bp.route('/')
@login_required
def list_integrations():
    configs = IntegrationConfig.query.order_by(IntegrationConfig.created_at.desc()).all()
    pending_count = Bom.query.filter_by(bom_type='MBOM', status='released', sync_status='not_synced').count()
    recent_logs = SyncLog.query.order_by(SyncLog.created_at.desc()).limit(30).all()
    return render_template('integrations/list.html', configs=configs, pending_count=pending_count, recent_logs=recent_logs)


@integration_bp.route('/<int:id>/update', methods=['POST'])
@login_required
def update_config(id):
    """更新集成配置（主要是 api_url）"""
    if current_user.role != 'admin':
        flash('只有管理员可以修改集成配置', 'danger')
        return redirect(url_for('integrations.list_integrations'))
    cfg = IntegrationConfig.query.get_or_404(id)
    cfg.api_url = request.form.get('api_url', cfg.api_url).strip()
    if request.form.get('db_name'):
        cfg.db_name = request.form.get('db_name').strip()
    if request.form.get('username'):
        cfg.username = request.form.get('username').strip()
    if request.form.get('api_key'):
        cfg.api_key = request.form.get('api_key').strip()
    db.session.commit()
    flash(f'已更新配置：{cfg.name}', 'success')
    return redirect(url_for('integrations.list_integrations'))


@integration_bp.route('/<int:id>/test', methods=['POST'])
@login_required
def test_config(id):
    """测试 Odoo 连接"""
    import xmlrpc.client, ssl, socket
    cfg = IntegrationConfig.query.get_or_404(id)
    socket.setdefaulttimeout(8)
    try:
        ctx = ssl.create_default_context()
        if os.environ.get('PLM_ODOO_INSECURE_SSL') == '1':
            ctx = ssl._create_unverified_context()
        common = xmlrpc.client.ServerProxy(f'{cfg.api_url}/xmlrpc/2/common', context=ctx, allow_none=True)
        v = common.version()
        uid = common.authenticate(cfg.db_name, cfg.username, cfg.api_key, {})
        if uid:
            return {'ok': True, 'version': v.get('server_version', '?'), 'uid': uid}
        return {'ok': False, 'error': '认证失败（uid 为空）'}
    except Exception as e:
        return {'ok': False, 'error': str(e)[:200]}


@integration_bp.route('/<int:id>/push-all', methods=['POST'])
@login_required
def push_all_mbom(id):
    """批量推送所有待同步的 mBOM 到 Odoo"""
    require_role('admin', 'manager')
    cfg = IntegrationConfig.query.get_or_404(id)
    if not cfg.is_active:
        flash('此集成配置未启用', 'warning')
        return redirect(url_for('integrations.list_integrations'))
    pending = Bom.query.filter_by(bom_type='MBOM', status='released', sync_status='not_synced').all()
    if not pending:
        flash('没有待推送的 mBOM', 'info')
        return redirect(url_for('boms.list_boms'))
    import subprocess, sys
    _base_dir = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
    runner = os.path.join(_base_dir, "push_to_odoo_runner.py")
    log_file = os.path.join(_base_dir, "plm_push.log")
    db_path = os.path.join(_base_dir, 'plm.db')
    pushed = 0
    for bom in pending:
        # 原子 CAS 抢占推送权
        affected = Bom.query.filter_by(id=bom.id).filter(
            Bom.sync_status == 'not_synced'
        ).update({'sync_status': 'pushing', 'sync_message': 'batch', 'sync_time': datetime.now()},
                 synchronize_session='fetch')
        db.session.commit()
        if affected == 0:
            continue
        subprocess.Popen([sys.executable, runner, str(bom.id), str(cfg.id), log_file, db_path],
                         stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
        pushed += 1
    if pushed:
        flash(f'已派发 {pushed} 个 mBOM 推送任务，1-2 分钟后刷新查看结果', 'info')
    else:
        flash('没有可推送的 mBOM（可能已被其他请求抢占）', 'info')
    return redirect(url_for('boms.list_boms', sync='not_synced'))


@integration_bp.route('/<int:id>/delete-config', methods=['POST'])
@login_required
def delete_config(id):
    """删除集成配置"""
    if current_user.role != 'admin':
        flash('只有管理员可以删除配置', 'danger')
        return redirect(url_for('integrations.list_integrations'))
    cfg = IntegrationConfig.query.get_or_404(id)
    db.session.delete(cfg)
    db.session.commit()
    flash(f'已删除配置：{cfg.name}', 'success')
    return redirect(url_for('integrations.list_integrations'))

# ── 工时管理 ──
@workhour_bp.route('/')
@login_required
def list_work_hours():
    hours = []
    return render_template('work_hours/list.html', hours=hours)

