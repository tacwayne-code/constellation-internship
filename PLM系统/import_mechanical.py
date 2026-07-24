#!/usr/bin/env python
"""
机械图批量导入 PLM 系统
扫描 Desktop/机械图 目录，自动创建：
  1. Product（产品结构）—— 用规格型号做唯一键，自动去重
  2. Document（图文档案）—— 图纸文件分类归档
  3. eBOM —— Excel 清单自动转为工程 BOM

策略：
  - 规格型号相同 → 复用同一个 Product（不重复建）
  - "用量"字段 → BOM 单台用量（不是批量采购数量）
  - 机号（710/910/812/...）→ Level 0 整机 Product
  - 每个 Excel 文件 → 一个 eBOM
"""

import os, sys, re, uuid, shutil
from datetime import datetime
from collections import defaultdict

# ── 配置 ──────────────────────────────
MECH_DIR = r'C:\Users\15897\Desktop\机械图'
SKIP_PATTERNS = ['~$', '.tmp']  # 跳过临时文件

# ── 初始化 Flask ──────────────────────
sys.path.insert(0, os.path.dirname(__file__))
from app import create_app
from app.models import db, User, Product, Bom, BomItem, Document, DocumentCategory, BomConversion

app = create_app()
ctx = app.app_context()
ctx.push()

admin = User.query.filter_by(username='admin').first()
if not admin:
    print('[ERROR] 请先用 admin/admin123 登录过系统')
    sys.exit(1)

# ── 全局计数器 ─────────────────────────
stats = {
    'products_created': 0, 'products_reused': 0,
    'boms_created': 0, 'bom_items': 0,
    'documents_created': 0, 'dirs_skipped': 0,
    'errors': []
}

# 已存在的产品缓存（规格型号 → Product）
existing_products = {}
for p in Product.query.all():
    existing_products[p.code] = p

def get_or_create_product(code, name, level=3):
    """用编码查重，没有就建新的 Level 3 零件"""
    # 清理编码
    code = str(code).strip().replace('\n', '').replace('\r', '')
    if not code or len(code) < 2:
        return None

    if code in existing_products:
        stats['products_reused'] += 1
        return existing_products[code]

    # 新建产品
    name = str(name).strip()[:200] if name else code
    # 根据名称判断层级
    if '整机' in name or '分光机' in name or '编带机' in name or '点胶机' in name or '贴片机' in name:
        level = 0  # 整机
    elif '组' in name or '总成' in name or '模组' in name or '模' in name:
        level = 1  # 部件/模组

    p = Product(code=code[:80], name=name, level=level, status='active')
    db.session.add(p)
    db.session.flush()
    existing_products[code] = p
    stats['products_created'] += 1
    return p


def create_machine_product(machine_name, machine_code):
    """创建整机 Product（如果不存在）"""
    if machine_code in existing_products:
        return existing_products[machine_code]
    p = Product(code=machine_code, name=machine_name, level=0, status='active')
    db.session.add(p)
    db.session.flush()
    existing_products[machine_code] = p
    stats['products_created'] += 1
    return p


# ── 1. 解析 Excel BOM 并创建 eBOM ─────────────
def parse_excel_bom(filepath):
    """读取 Excel 文件（支持 .xls 和 .xlsx），提取所有物料行"""
    ext = filepath.rsplit('.', 1)[-1].lower() if '.' in filepath else ''

    if ext == 'xls':
        import xlrd
        try:
            wb = xlrd.open_workbook(filepath)
        except Exception as e:
            stats['errors'].append(f'无法打开 {filepath}: {e}')
            return []
        sheets = [(sn, 'xls', wb.sheet_by_name(sn)) for sn in wb.sheet_names()]
    else:
        import openpyxl
        try:
            wb = openpyxl.load_workbook(filepath, data_only=True)
        except Exception as e:
            stats['errors'].append(f'无法打开 {filepath}: {e}')
            return []
        sheets = [(sn, 'xlsx', wb[sn]) for sn in wb.sheetnames]

    def cell(sh, r, c, fmt='xls'):
        try:
            if fmt == 'xls':
                return str(sh.cell_value(r, c)).strip()
            else:
                v = sh.cell(row=r+1, column=c+1).value
                return str(v).strip() if v is not None else ''
        except:
            return ''

    items = []
    for sn, fmt, sh in sheets:
        if fmt == 'xls':
            nrows, ncols = sh.nrows, sh.ncols
        else:
            nrows, ncols = sh.max_row, sh.max_column
        # 找表头行（含"序号"的行）
        header_row = None
        for r in range(min(nrows, 30)):
            row_vals = [cell(sh, r, c, fmt) for c in range(min(ncols, 16))]
            if any('序号' in v for v in row_vals):
                header_row = r
                break

        if header_row is None:
            continue

        # 解析表头列索引
        header = [cell(sh, header_row, c, fmt) for c in range(ncols)]
        col_map = {}
        for i, h in enumerate(header):
            if '序号' in h: col_map['seq'] = i
            elif '名称' in h: col_map['name'] = i
            elif '规格型号' in h or '型号' in h or '规格' in h: col_map['spec'] = i
            elif '品牌' in h: col_map['brand'] = i
            elif '用量' in h: col_map['usage'] = i
            elif '数量' in h and '用量' not in h: col_map['qty'] = i
            elif '单位' in h: col_map['unit'] = i
            elif '供应商' in h: col_map['supplier'] = i
            elif '单价' in h or '价格' in h: col_map['price'] = i
            elif '要求' in h or '备注' in h: col_map['note'] = i

        if 'name' not in col_map and 'spec' not in col_map:
            continue

        # 提取数据行
        for r in range(header_row + 1, nrows):
            seq_val = cell(sh, r, col_map.get('seq', 0), fmt)
            name_val = cell(sh, r, col_map.get('name', 1), fmt)
            spec_val = cell(sh, r, col_map.get('spec', 2), fmt)

            if not seq_val or seq_val == '' or seq_val == '0.0' or seq_val == '0':
                continue

            try:
                float(seq_val)
            except ValueError:
                continue

            if not name_val and not spec_val:
                continue

            # 获取用量
            usage = 1.0
            if 'usage' in col_map:
                try:
                    usage = float(cell(sh, r, col_map['usage'], fmt))
                except (ValueError, TypeError):
                    usage = 1.0
            elif 'qty' in col_map:
                try:
                    usage = float(cell(sh, r, col_map['qty'], fmt))
                except (ValueError, TypeError):
                    usage = 1.0

            unit = cell(sh, r, col_map['unit'], fmt) if 'unit' in col_map else '个'
            if not unit or unit == '':
                unit = '个'

            brand = cell(sh, r, col_map['brand'], fmt) if 'brand' in col_map else ''
            supplier = cell(sh, r, col_map['supplier'], fmt) if 'supplier' in col_map else ''
            note = cell(sh, r, col_map.get('note', -1), fmt) if 'note' in col_map else ''

            items.append({
                'name': name_val,
                'spec': spec_val,
                'brand': brand,
                'usage': usage,
                'unit': unit,
                'supplier': supplier,
                'note': note,
            })

    return items


def import_bom_from_excel(filepath, bom_name, product):
    """从 Excel 创建一个 eBOM"""
    items = parse_excel_bom(filepath)
    if not items:
        stats['errors'].append(f'{filepath}: 没有可用的物料数据')
        return None

    # 去重合并：同一规格型号合并用量
    merged = {}
    for item in items:
        key = item['spec'] if item['spec'] else item['name']
        if key in merged:
            merged[key]['usage'] += item['usage']
        else:
            merged[key] = item

    # 创建 eBOM
    bom = Bom(
        name=bom_name,
        description=f'从 Excel 导入：{os.path.basename(filepath)}',
        bom_type='EBOM',
        product_id=product.id if product else None,
        created_by=admin.id,
        status='draft'
    )
    db.session.add(bom)
    db.session.flush()

    # 创建 BomItem（每个物料行）
    seq = 0
    for key, item in merged.items():
        # 创建或复用 Product
        spec = item['spec'] if item['spec'] else item['name']
        p = get_or_create_product(spec, item['name'], level=3)
        if not p:
            continue

        bi = BomItem(
            bom_id=bom.id,
            product_id=p.id,
            quantity=item['usage'],
            unit=item['unit'],
            seq=seq + 1,
            note=f"{item['brand']} | {item['supplier']} | {item['note']}"[:200]
        )
        db.session.add(bi)
        seq += 1

    stats['boms_created'] += 1
    stats['bom_items'] += seq
    return bom


# ── 2. 图文档案导入 ─────────────────────
def get_or_create_category(name, parent_id=None):
    """创建文档分类"""
    existing = DocumentCategory.query.filter_by(name=name, parent_id=parent_id).first()
    if existing:
        return existing
    cat = DocumentCategory(name=name, parent_id=parent_id)
    db.session.add(cat)
    db.session.flush()
    return cat


def import_document(filepath, rel_path, category_map):
    """导入单个文档，基于目录层级分类"""
    if not os.path.isfile(filepath):
        return

    fname = os.path.basename(filepath)
    # 跳过临时文件
    if any(p in fname for p in SKIP_PATTERNS):
        return

    size = os.path.getsize(filepath)
    if size == 0:
        return

    # 判断文件类型
    ext = fname.rsplit('.', 1)[-1].lower() if '.' in fname else ''
    doc_type_map = {
        'pdf': 'PDF图纸', 'dwg': 'CAD图纸', 'dxf': 'DXF图纸',
        'slddrw': 'SolidWorks工程图', 'sldprt': 'SolidWorks零件', 'sldasm': 'SolidWorks装配',
        'step': 'STEP3D', 'stp': 'STEP3D', 'xt': 'XT3D',
        'doc': 'Word文档', 'docx': 'Word文档',
        'xls': 'Excel表格', 'xlsx': 'Excel表格',
        'jpg': '图片', 'png': '图片', 'jpeg': '图片',
        'zip': '压缩包', 'rar': '压缩包',
    }
    doc_type = doc_type_map.get(ext, f'其他({ext})')

    # 解析目录层级，确定分类
    parts = rel_path.replace('\\', '/').split('/')
    category = None

    # 根据路径中的关键词分类
    if any('五金' in p for p in parts):
        category = category_map.get('五金件加工图纸')
    elif any('钣金' in p for p in parts):
        category = category_map.get('钣金加工图纸')
    elif any('标准件' in p for p in parts):
        category = category_map.get('标准件图纸')
    elif any('清单' in p for p in parts) or any('BOM' in p for p in parts):
        category = category_map.get('BOM清单')
    elif any('CAD' in p for p in parts):
        category = category_map.get('CAD图纸')
    elif any('PDF' in p for p in parts):
        category = category_map.get('PDF图纸')
    elif any('3D' in p for p in parts) or any('STEP' in p.upper() for p in parts) or any('XT' in p.upper() for p in parts):
        category = category_map.get('3D模型')

    if not category:
        category = category_map.get('其他图纸')

    # 创建文档记录（不复制文件，只记录路径引用）
    doc = Document(
        title=fname.replace('.' + ext, ''),
        description=f'自动导入 | 路径: {rel_path} | 类型: {doc_type} | 大小: {size} bytes',
        category_id=category.id if category else None,
        tags=','.join([doc_type] + [p for p in parts if p][:3]),
        status='published',
        author_id=admin.id,
        file_name=fname,
        file_size=size
    )
    db.session.add(doc)
    stats['documents_created'] += 1
    return doc


# ── 3. 主流程 ───────────────────────────
def main():
    print('=' * 60)
    print('  机械图批量导入 PLM 系统')
    print(f'  源目录: {MECH_DIR}')
    print(f'  目标: 自动创建产品/物料/eBOM/图文档案')
    print('=' * 60)

    if not os.path.exists(MECH_DIR):
        print(f'[ERROR] 目录不存在: {MECH_DIR}')
        return

    # ── 创建文档分类 ──
    print('\n[1/4] 初始化文档分类...')
    category_map = {
        '五金件加工图纸': get_or_create_category('五金件加工图纸'),
        '钣金加工图纸': get_or_create_category('钣金加工图纸'),
        '标准件图纸': get_or_create_category('标准件图纸'),
        'BOM清单': get_or_create_category('BOM清单'),
        'CAD图纸': get_or_create_category('CAD图纸'),
        'PDF图纸': get_or_create_category('PDF图纸'),
        '3D模型': get_or_create_category('3D模型'),
        '其他图纸': get_or_create_category('其他图纸'),
    }
    print(f'  文档分类: {len(category_map)} 个')

    # ── 扫目录，导入 BOM 和文档 ──
    print('\n[2/4] 扫描 Excel BOM 并创建 eBOM...')
    excel_count = 0

    # 遍历所有 Excel（先做 BOM）
    for root, dirs, files in os.walk(MECH_DIR):
        for f in files:
            if any(p in f for p in SKIP_PATTERNS):
                continue
            ext = f.rsplit('.', 1)[-1].lower() if '.' in f else ''
            if ext not in ('xls', 'xlsx'):
                continue

            filepath = os.path.join(root, f)
            rel_path = os.path.relpath(filepath, MECH_DIR)

            # 提取机号作为产品名
            # 匹配模式: 710分光机, 812点胶机, 搅拌机, 固晶机, 贴片机
            parts = rel_path.replace('\\', '/').split('/')
            machine_name = ''
            for p in parts:
                m = re.search(r'(\d{3})\s*(分光|编带|点胶|贴片|搅拌|固晶|上下板|打码)', p)
                if m:
                    machine_name = m.group(0).strip()
                    break
                # 匹配无编号的机型名
                m2 = re.search(r'(搅拌机|固晶机|贴片机|上下板机|打码机|编带机|分光机|点胶机)', p)
                if m2:
                    machine_name = m2.group(0).strip()
                    break

            if not machine_name:
                # 从文件名提取
                m2 = re.search(r'(\d{3})\s*(分光|编带|点胶|贴片|搅拌)', f)
                if m2:
                    machine_name = m2.group(0).strip()

            if not machine_name:
                stats['dirs_skipped'] += 1
                continue

            # 创建整机 Product
            machine_code = re.sub(r'\s+', '', machine_name)
            machine = create_machine_product(machine_name, machine_code)

            # 提取版本信息
            version = 'V1'
            for p in parts:
                vm = re.search(r'(V\d+)', p, re.IGNORECASE)
                if vm:
                    version = vm.group(1).upper()
                    break
                if '修改' in p:
                    vm2 = re.search(r'(\d{4}\.\d{1,2}\.\d{1,2})', p)
                    if vm2:
                        version = vm2.group(1)

            bom_name = f'{machine_name} {version} 配件清单'
            bom = import_bom_from_excel(filepath, bom_name, machine)
            excel_count += 1

            if excel_count % 10 == 0:
                print(f'  已处理 {excel_count} 个 Excel...')

            # 每 20 个 Excel commit 一次
            if excel_count % 20 == 0:
                db.session.commit()
                print(f'  [commit] products={stats["products_created"]}/{stats["products_reused"]}, boms={stats["boms_created"]}')

    db.session.commit()
    print(f'  完成! {excel_count} 个 Excel, {stats["boms_created"]} 个 eBOM, {stats["bom_items"]} 条物料')

    # ── 导入图文档案（大文件批量，采样子集）──
    print('\n[3/4] 导入图文档案（图纸文件）...')
    doc_count = 0
    doc_limit = 5000  # 限制数量，避免 SQLite 爆炸
    doc_imported = 0

    for root, dirs, files in os.walk(MECH_DIR):
        for f in files:
            if any(p in f for p in SKIP_PATTERNS):
                continue
            ext = f.rsplit('.', 1)[-1].lower() if '.' in f else ''

            # 重点是图纸、PDF、CAD 文件
            valid_exts = {'pdf', 'dwg', 'dxf', 'slddrw', 'step', 'stp'}
            if ext not in valid_exts:
                continue

            filepath = os.path.join(root, f)
            rel_path = os.path.relpath(filepath, MECH_DIR)
            import_document(filepath, rel_path, category_map)
            doc_imported += 1

            if doc_imported >= doc_limit:
                break

            if doc_imported % 500 == 0:
                print(f'  已导入 {doc_imported} 个图纸文件...')
                db.session.commit()

        if doc_imported >= doc_limit:
            break

    db.session.commit()
    print(f'  完成! 导入 {doc_imported} 个图纸文件')

    # ── 统计报告 ──
    print('\n[4/4] 导入完成! 统计报告:')
    print(f'  产品（物料）: 新建 {stats["products_created"]}, 复用 {stats["products_reused"]}')
    print(f'  eBOM: {stats["boms_created"]} 个')
    print(f'  BOM 明细: {stats["bom_items"]} 条')
    print(f'  图文档案: {doc_imported} 个')
    print(f'  跳过: {stats["dirs_skipped"]} 个无法识别机号的 Excel')

    if stats['errors']:
        print(f'\n  错误 ({len(stats["errors"])}):')
        for e in stats['errors'][:10]:
            print(f'    - {e[:120]}')

    print('\n=== 导入完成！现在可以在 PLM 中查看数据 ===')
    print(f'  访问: http://localhost:5000')
    print(f'  BOM 管理 → 查看 eBOM → 结构管理 → 查看产品')


if __name__ == '__main__':
    main()
