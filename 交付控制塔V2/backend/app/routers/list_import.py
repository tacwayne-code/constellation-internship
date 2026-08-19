"""清单导入 API：解析 → 识别 → 推荐 → 批量建单"""
from __future__ import annotations

import csv
import io
import logging
import re
from typing import Annotated

from fastapi import APIRouter, Depends, File, Query, Response, UploadFile
from pydantic import BaseModel, Field

from app.config import get_settings
from app.services.import_matching import (
    batch_create_purchase_orders,
    load_product_index,
    match_products,
    recommend_suppliers,
)
from app.services.odoo.client import OdooClient

logger = logging.getLogger(__name__)
router = APIRouter(prefix="/api/procurement/list", tags=["procurement-list-import"])


def get_client() -> OdooClient:
    return OdooClient.get_instance(get_settings())


ClientDep = Annotated[OdooClient, Depends(get_client)]


# ---- 请求模型 ----
class ListRow(BaseModel):
    name: str = Field(..., description="物料名称，如「平垫圈16×3」")
    qty: float = Field(1, description="数量")
    spec: str = Field("", description="可选，手工指定规格")
    code: str = Field("", description="可选，物料编码")
    supplier: str = Field("", description="可选，清单自带供应商")
    remark: str = Field("", description="可选，清单备注")


class ParseRequest(BaseModel):
    text: str = Field(..., description="粘贴的清单文本，每行一条，支持 逗号/制表符/空格 分隔「名称,数量」")


class MatchRequest(BaseModel):
    rows: list[ListRow]


class RecommendRequest(BaseModel):
    product_ids: list[int] = Field(..., description="product.template id 列表")


class CreatePoLine(BaseModel):
    """采购行：product_id 与 name 二选一（无 product_id 时按 name 自动建料）"""
    product_id: int | None = Field(None, description="已识别产品的 product.template id；缺省时按 name 自动建料")
    name: str | None = Field(None, description="物料名称（无 product_id 时必填，按名称自动找/建产品）")
    qty: float = Field(..., gt=0)
    partner_id: int | None = None
    supplier_name: str | None = Field(None, description="供应商名称（无 partner_id 时按名称自动找/建供应商）")
    price: float | None = None
    delay: int | None = None
    remark: str | None = Field(None, description="可选，清单备注（写入采购行 note）")
    code: str | None = Field(None, description="可选，清单「编号」列；与 name 拼接写入采购行 name")


class CreatePoRequest(BaseModel):
    lines: list[CreatePoLine]
    auto_create_product: bool = Field(
        False,
        description="True=按名称自动建产品主数据；False（默认）=不建料，无 product_id 的行挂「临时外购件」占位产品",
    )
    urgent: bool = Field(
        True,
        description="True=标记为紧急采购单（priority=1，进紧急采购看板）；False=普通采购单",
    )
    purchase_date: str | None = Field(
        None,
        description="采购时间（订单时间），写入 purchase.order.date_order；支持 'YYYY-MM-DD' 或 'YYYY-MM-DD HH:MM'；缺省用 Odoo 当前时间",
    )
    delivery_date: str | None = Field(
        None,
        description="交货时间（计划到货时间），写入订单行 date_planned；支持 'YYYY-MM-DD' 或 'YYYY-MM-DD HH:MM'；缺省按供应商交期（今天+delay）",
    )


# ---- 接口 ----

@router.post("/parse")
async def parse_list(req: ParseRequest):
    """粘贴清单文本 → 解析出行（名称 + 数量）。

    支持逗号 / 制表符 / 空格分隔，最后一列若为数字则当作数量。
    兼容全角逗号「，」与全角制表符（中文输入法默认输出全角）。
    """
    rows: list[dict] = []
    for raw_line in req.text.splitlines():
        line = raw_line.strip()
        if not line or line.startswith("#"):
            continue
        # 判定分隔符（兼容全角逗号 / 全角制表符）
        if "\t" in line or "\u3000" in line:
            delim = "\t" if "\t" in line else "\u3000"
        elif "," in line or "，" in line:
            delim = "," if "," in line else "，"
        else:
            delim = None
        if delim:
            parts = [p.strip() for p in line.split(delim) if p.strip()]
        else:
            parts = re.split(r"\s{2,}|\s+", line.strip())
            parts = [p for p in parts if p]

        if not parts:
            continue
        # 末尾若是数量（支持「48」「48个」「48+3」），则拆为数量；否则当作名称的一部分
        qty = _parse_qty(parts[-1])
        if qty is not None:
            name = " ".join(parts[:-1])
        else:
            qty = 1.0
            name = " ".join(parts)
        if name:
            rows.append({"name": name, "qty": qty})

    return {"count": len(rows), "rows": rows}


# 数量单位后缀（如「48个」「3.5件」「2台套」），解析时去掉后转数字
_UNIT_RE = re.compile(
    r"(?:个|件|套|台|只|片|根|块|条|米|桶|箱|包|卷|张|支|对|颗|把|pcs?|sets?|ea|units?|台套|套件|组|批)+\.?$",
    re.IGNORECASE,
)


def _parse_qty(value) -> float | None:
    """解析数量：48+3 → 51；48 → 48；'48个' → 48；'48 + 3件' → 51；非数字 → None（视为分类行）"""
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    s = str(value).strip().replace("＋", "+").replace(" ", "").replace("　", "")
    if not s:
        return None
    s = _UNIT_RE.sub("", s)  # 去掉末尾单位（48件→48，48个→48，2台套→2）
    if not s:
        return None
    if "+" in s:
        try:
            return sum(float(x) for x in s.split("+") if x)
        except ValueError:
            return None
    try:
        return float(s)
    except ValueError:
        return None


def _norm_header(value) -> str:
    """表头归一化：去空格 + 全角括号转半角 + 去掉括号注释。

    使「数量（个）」「数量 (个)」都能识别为「数量」列。
    """
    if value is None:
        return ""
    s = str(value).strip().replace("　", " ").replace(" ", "")
    s = s.replace("（", "(").replace("）", ")")
    return re.sub(r"\([^)]*\)", "", s)


def _parse_xlsx(content: bytes) -> list[dict]:
    """解析 .xlsx：自动探测「名称/数量/编号/供应商/备注」列，跳过分类行/空行，数量支持 48+3 求和。"""
    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    ws = wb.worksheets[0]

    name_col = 2  # 默认 B 列
    qty_col = 4   # 默认 D 列
    code_col = 3  # 默认 C 列（编号）
    supplier_col = 7  # 默认 G 列（供应商）
    remark_col = 10   # 默认 J 列（备注）
    header_row = 0
    # 前 15 行探测表头：同一行同时含「名称」「数量」列才认定（精确匹配，避免「设备名称」误判）
    for r in range(1, min(ws.max_row, 15) + 1):
        found_name = found_qty = found_code = found_supplier = found_remark = None
        for c in range(1, min(ws.max_column, 30) + 1):
            v = ws.cell(row=r, column=c).value
            if v is None:
                continue
            text = _norm_header(v)
            if found_name is None and text in ("名称", "品名", "物料", "物料名称", "名称/规格"):
                found_name = c
            if found_qty is None and text in ("数量", "用量", "需求数量", "采购数量"):
                found_qty = c
            if found_code is None and text in ("编号", "物料编码", "代码", "编码", "default_code"):
                found_code = c
            if found_supplier is None and text in ("供应商", "供用商", "供应商名称", "供应商名", "厂商", "厂家"):
                found_supplier = c
            if found_remark is None and text in ("备注", "说明", "描述", "remark"):
                found_remark = c
        if found_name and found_qty:
            name_col, qty_col, header_row = found_name, found_qty, r
            if found_code is not None:
                code_col = found_code
            if found_supplier is not None:
                supplier_col = found_supplier
            if found_remark is not None:
                remark_col = found_remark
            break

    rows: list[dict] = []
    for r in range(header_row + 1, ws.max_row + 1):
        name = ws.cell(row=r, column=name_col).value
        if name is None:
            continue
        name = str(name).strip()
        if not name or name in ("名称", "品名", "物料"):
            continue
        qty = _parse_qty(ws.cell(row=r, column=qty_col).value)
        if qty is None:
            continue  # 数量为空 → 分类行（如「天轨用」「地轨用」）
        code_val = ws.cell(row=r, column=code_col).value
        code = str(code_val).strip() if code_val is not None and str(code_val).strip() else ""
        supplier_val = ws.cell(row=r, column=supplier_col).value
        supplier = str(supplier_val).strip() if supplier_val is not None and str(supplier_val).strip() else ""
        remark_val = ws.cell(row=r, column=remark_col).value
        remark = str(remark_val).strip() if remark_val is not None and str(remark_val).strip() else ""
        rows.append({"name": name, "qty": qty, "code": code, "supplier": supplier, "remark": remark})
    return rows


def _parse_csv_text(text: str) -> list[dict]:
    """解析 CSV 文本（含表头探测）：定位名称/数量/编号/供应商/备注列。"""
    reader = csv.reader(io.StringIO(text))
    data = list(reader)
    if not data:
        return []
    # 找表头（精确匹配，同一行含名称+数量）
    header_idx = -1
    name_col = qty_col = -1
    code_col = -1
    supplier_col = -1
    remark_col = -1
    for i, row in enumerate(data[:15]):
        fn = fq = fc = fs = fr = -1
        for j, cell in enumerate(row):
            c = _norm_header(cell)
            if fn < 0 and c in ("名称", "品名", "物料", "物料名称"):
                fn = j
            if fq < 0 and c in ("数量", "用量", "需求数量"):
                fq = j
            if fc < 0 and c in ("编号", "物料编码", "代码", "编码", "default_code"):
                fc = j
            if fs < 0 and c in ("供应商", "供用商", "供应商名称", "供应商名", "厂商", "厂家"):
                fs = j
            if fr < 0 and c in ("备注", "说明", "描述", "remark"):
                fr = j
        if fn >= 0 and fq >= 0:
            name_col, qty_col, header_idx = fn, fq, i
            if fc >= 0:
                code_col = fc
            if fs >= 0:
                supplier_col = fs
            if fr >= 0:
                remark_col = fr
            break
    if name_col < 0:
        name_col = 0
    if qty_col < 0:
        qty_col = 1
    if code_col < 0:
        code_col = -1
    if supplier_col < 0:
        supplier_col = -1
    if remark_col < 0:
        remark_col = -1
    rows = []
    for row in data[header_idx + 1:]:
        if name_col >= len(row):
            continue
        name = str(row[name_col]).strip()
        if not name:
            continue
        qty_raw = row[qty_col] if qty_col < len(row) else ""
        qty = _parse_qty(qty_raw)
        if qty is None:
            qty = 1.0
        code = str(row[code_col]).strip() if (code_col >= 0 and code_col < len(row)) else ""
        supplier = str(row[supplier_col]).strip() if (supplier_col >= 0 and supplier_col < len(row)) else ""
        remark = str(row[remark_col]).strip() if (remark_col >= 0 and remark_col < len(row)) else ""
        rows.append({"name": name, "qty": qty, "code": code, "supplier": supplier, "remark": remark})
    return rows


@router.post("/upload")
async def upload_list(
    file: UploadFile = File(..., description="Excel(.xlsx) 或 CSV 文件"),
):
    """上传清单文件（.xlsx/.csv）→ 解析出行（名称 + 数量）。

    自动探测「名称」「数量」列，跳过分类行；数量支持「48+3」求和。
    """
    try:
        content = await file.read()
        filename = (file.filename or "").lower()
        if filename.endswith(".csv"):
            rows = _parse_csv_text(content.decode("utf-8-sig", errors="ignore"))
        else:
            rows = _parse_xlsx(content)
    except Exception as e:  # noqa: BLE001
        logger.exception("upload_list failed")
        return {"ok": False, "error": f"文件解析失败: {str(e)[:160]}"}
    return {"ok": True, "count": len(rows), "rows": rows}


@router.post("/match")
async def match_list(
    req: MatchRequest,
    client: ClientDep,
    resp: Response,
):
    """识别配件：清单行 → 匹配 product.template（类型词 + 规格）。

    返回每行的 matched/product_id/score/candidates/action。
    """
    try:
        products = await load_product_index(client)
        results = await match_products(client, [r.model_dump() for r in req.rows], products=products)
    except Exception as e:  # noqa: BLE001
        logger.exception("match_list failed")
        return {"ok": False, "error": str(e)[:160]}

    auto = sum(1 for r in results if r["action"] == "auto")
    choose = sum(1 for r in results if r["action"] == "choose")
    create = sum(1 for r in results if r["action"] == "create")
    resp.headers["X-Data-Source"] = "odoo"
    return {
        "ok": True,
        "stats": {"total": len(results), "auto": auto, "choose": choose, "create": create},
        "rows": results,
    }


@router.post("/recommend")
async def recommend_list(
    req: RecommendRequest,
    client: ClientDep,
    resp: Response,
):
    """推荐供应商：product.template id 列表 → 每个产品的供应商 Top5（评分降序）。"""
    try:
        data = await recommend_suppliers(client, req.product_ids)
    except Exception as e:  # noqa: BLE001
        logger.exception("recommend_list failed")
        return {"ok": False, "error": str(e)[:160]}
    resp.headers["X-Data-Source"] = "odoo"
    return {"ok": True, "suppliers": data}


@router.post("/create-po")
async def create_po_list(
    req: CreatePoRequest,
    client: ClientDep,
    resp: Response,
):
    """批量建采购单：按供应商聚合，priority=1 紧急 + date_planned。

    auto_create_product=False（默认）：无 product_id 的行不建料，挂「临时外购件」占位产品，
    采购行描述写实际物料名，采购单直接进入后续环节。
    """
    try:
        data = await batch_create_purchase_orders(
            client,
            [l.model_dump() for l in req.lines],
            auto_create_product=req.auto_create_product,
            urgent=req.urgent,
            purchase_date=req.purchase_date,
            delivery_date=req.delivery_date,
        )
    except Exception as e:  # noqa: BLE001
        logger.exception("create_po_list failed")
        return {"ok": False, "error": str(e)[:160]}
    resp.headers["X-Data-Source"] = "odoo"
    return {"ok": True, **data}


@router.get("/partners")
async def list_partners(
    client: ClientDep,
    resp: Response,
    limit: int = Query(100, ge=1, le=500),
):
    """常用供应商列表（按 supplier_rank 降序），供清单导入件手工指定供应商。"""
    try:
        partners = await client.search_read(
            "res.partner",
            [["supplier_rank", ">", 0]],
            ["id", "name", "supplier_rank"], limit=limit, order="supplier_rank desc, name",
        )
    except Exception as e:  # noqa: BLE001
        logger.exception("list_partners failed")
        return {"ok": False, "error": str(e)[:160]}
    resp.headers["X-Data-Source"] = "odoo"
    return {"ok": True, "partners": [
        {"partner_id": p["id"], "name": p.get("name") or "", "supplier_rank": p.get("supplier_rank") or 0}
        for p in partners
    ]}


class PartnerCreate(BaseModel):
    name: str = Field(..., min_length=1, description="新供应商名称")


@router.post("/partners")
async def create_partner(
    req: PartnerCreate,
    client: ClientDep,
    resp: Response,
):
    """新建供应商（res.partner：公司 + 供应商标记），供「新供应商」场景使用。"""
    name = req.name.strip()
    if not name:
        return {"ok": False, "error": "供应商名称不能为空"}
    try:
        pid = await client.create("res.partner", {
            "name": name,
            "is_company": True,
        })
        # Odoo 18 的 supplier_rank 在 create 时会被忽略，需 write 补设
        try:
            await client.write("res.partner", [pid], {"supplier_rank": 1})
        except Exception:  # noqa: BLE001
            pass
    except Exception as e:  # noqa: BLE001
        logger.exception("create_partner failed")
        return {"ok": False, "error": str(e)[:160]}
    resp.headers["X-Data-Source"] = "odoo"
    return {"ok": True, "partner_id": pid, "name": name}
